# -*- coding: utf-8 -*-
"""
    @Time : 2024-04-14
    Description : 有用
"""
import pandas as pd
import requests
from bs4 import BeautifulSoup

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

from datetime import datetime


def optim_excel(excel_file):
    # 打开现有的 Excel 文件
    wb = load_workbook(excel_file)
    ws = wb.active

    # 设置单元格样式：居中对齐和添加边框
    alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin'))

    # 遍历所有单元格应用样式
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = alignment
            cell.border = border

    # 保存修改后的工作簿
    wb.save(excel_file)
    print(f"美化单元格保存路径: {excel_file}")


# columns_list = ['编号', '赛事', '开赛时间', '主队', '客队', '胜', '胜变化', '平', '平变化', '负', '负变化',
#                 'wdl变化时间', '让胜', '让胜变化', '让平', '让平变化', '让负', '让负变化', 'h_wdl变化时间']
columns_list_p1 = ['编号', '赛事', '开赛时间', '主队', '客队']
columns_list_p2 = ['胜', '胜变化', '平', '平变化', '负', '负变化', 'wdl变化时间']
columns_list_p3 = ['让胜', '让胜变化', '让平', '让平变化', '让负', '让负变化', 'h_wdl变化时间']
df1 = pd.DataFrame(columns=columns_list_p1)
df2 = pd.DataFrame(columns=columns_list_p2)
df3 = pd.DataFrame(columns=columns_list_p3)
df_res = pd.DataFrame()

header = {
    "User-Agent": "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/"
                  "86.0.4240.198 Safari/537.36"
}

url = 'https://trade.500.com/jczq/?playid=269&g=2&date=2024-04-16'
print(f'url: {url}')

response = requests.get(url, headers=header)
content = response.content.decode('ISO-8859-1')
soup = BeautifulSoup(response.text, 'lxml')

# trs = soup.find_all('tr', class_="bet-tb-tr bet-tb-end")
trs = soup.find_all('tr', class_="bet-tb-tr")
for tr in trs:
    # df1 = df1.append(pd.Series([], dtype='float64'), ignore_index=True)

    home_team = tr['data-homesxname']
    away_team = tr['data-awaysxname']
    match_events = tr['data-simpleleague']
    match_date = tr['data-matchdate']
    match_time = tr['data-matchtime']
    match_datetime = match_date + ' ' + match_time
    interface_date = tr['data-processdate']
    match_no = tr['data-matchnum']
    data_id = tr['data-id']
    print(f'home_team: {home_team}')

    # 请求数据的 sp 变化接口
    print(f"zxid: {data_id}")
    print(f"date: {interface_date}")
    current_date = datetime.now().date()
    if interface_date == str(current_date):
        interface_date = ''
    win_draw_lose_url = 'https://trade.500.com/interface/request.php?q=public.readfile&step=readpl&playid=354&zxid={}&date={}'.format(
        data_id, interface_date)
    print(f"win_draw_lose_url: {win_draw_lose_url}")
    handicap_wdl_url = 'https://trade.500.com/interface/request.php?q=public.readfile&step=readpl&playid=269&zxid={}&date={}'.format(
        data_id, interface_date)
    # print("胜平负 sp 变化接口: {}".format(win_draw_lose_url))
    # print("受让球 sp 变化接口: {}".format(handicap_wdl_url))
    wdl_response = requests.get(win_draw_lose_url)
    handicap_wdl_response = requests.get(handicap_wdl_url)
    wdl_data = wdl_response.json()
    handicap_wdl_data = handicap_wdl_response.json()

    # print(len(wdl_data['list']))
    # print(len(handicap_wdl_data['list']))

    if len(wdl_data['list']) >= len(handicap_wdl_data['list']):
        for i in range(len(wdl_data['list'])):
            # print(i)
            data = {
                '编号': match_no,
                '赛事': match_events,
                '开赛时间': match_datetime,
                '主队': home_team,
                '客队': away_team
            }
            df1 = df1.append(data, ignore_index=True)

    else:
        for i in range(len(handicap_wdl_data['list'])):
            data = {
                '编号': match_no,
                '赛事': match_events,
                '开赛时间': match_datetime,
                '主队': home_team,
                '客队': away_team
            }
            df1 = df1.append(data, ignore_index=True)

    for i in wdl_data['list']:
        data = {
            '胜': i['win'],
            '胜变化': i['w'],
            '平': i['draw'],
            '平变化': i['d'],
            '负': i['lost'],
            '负变化': i['l'],
            'wdl变化时间': i['time']
        }
        df2 = df2.append(data, ignore_index=True)
    if len(wdl_data['list']) <= len(handicap_wdl_data['list']):
        for _ in range(len(handicap_wdl_data['list']) - len(wdl_data['list'])):
            df2 = df2.append(pd.Series([], dtype='float64'), ignore_index=True)

    for i in handicap_wdl_data['list']:
        data = {
            '让胜': i['win'],
            '让胜变化': i['w'],
            '让平': i['draw'],
            '让平变化': i['d'],
            '让负': i['lost'],
            '让负变化': i['l'],
            'h_wdl变化时间': i['time']
        }
        df3 = df3.append(data, ignore_index=True)
    if len(wdl_data['list']) >= len(handicap_wdl_data['list']):
        for _ in range(len(wdl_data['list']) - len(handicap_wdl_data['list'])):
            df3 = df3.append(pd.Series([], dtype='float64'), ignore_index=True)

    # print('')
    df1 = df1.append(pd.Series([], dtype='float64'), ignore_index=True)
    df2 = df2.append(pd.Series([], dtype='float64'), ignore_index=True)
    df3 = df3.append(pd.Series([], dtype='float64'), ignore_index=True)

df_res = pd.concat([df1, df2, df3], axis=1)
# print(df_res)

# 保存新文件
output_xlsx_file = 'test-04-16-2.xlsx'
df_res.to_excel(output_xlsx_file, index=False)

# 美化表格
optim_excel(output_xlsx_file)
