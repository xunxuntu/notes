"""
    从网站爬取某一天的数据，并进行分析，
    得到以年月日命名的excel 和 柱状图，
    e.g. proceed_2023-11-02.xlsx 和 proceed_2023-11-02.png
    查询过去三十天内的某一天的胜负情况
        有用
        传入一个日期参数即可运行
"""

import os
import requests
import time
from bs4 import BeautifulSoup
import pandas as pd
import xlwings as xw
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from huatu import show_it


def search_func(x_date):
    header = {
        "User-Agent": "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/"
                      "86.0.4240.198 Safari/537.36"
    }

    # columns = ['让球','胜','平','负']
    title = {'编号': '', '赛事': '', '开赛时间': '', '主队排名': '', '主队': '', '比分': '', '客队': '', '客队排名': '',
             '让球': ''}
    data = {'胜': '', '平': '', '负': ''}
    serial_numbers, game_names, game_times = [], [], []
    i0, i1, a0, a1, a2 = [], [], [], [], []
    concede, victory, flat, defeat = [], [], [], []
    victory_index, flat_index, defeat_index = [], [], []

    tm_rng = pd.date_range(start=x_date, end=x_date, freq='D')
    print(f'tm_rng: {tm_rng}')

    for tms in tm_rng:
        tm = tms.strftime('%Y-%m-%d')
        url = 'https://trade.500.com/jczq/?date={}'.format(tm)
        print(f'tm: {tm}')
        print(f'url: {url}')

        respone = requests.get(url, headers=header)
        print(respone.status_code)
        print(f'respone: {respone}')

        content = respone.content.decode('ISO-8859-1')
        # print(f'content: {content}')

        soup = BeautifulSoup(respone.text, 'lxml')
        trs = soup.find_all('tr', class_="bet-tb-tr bet-tb-end")
        # print(f'soup: {soup}')

        print('=' * 90)
        print(f'trs: {trs}')
        print('=' * 90)

        for tr in trs:
            print('*' * 90)
            print(f'tr: {tr}')

            # 编号
            td_no = tr.find_all('td', class_="td td-no")
            print(f'td_no: {td_no}')

            serial_number = td_no[0].text
            print(f'serial_number: {serial_number}')

            # 重复赋值是为了合并之后能够进行筛选
            serial_numbers.extend((serial_number, serial_number))
            # 赛事
            td_evt = tr.find_all('td', class_="td td-evt")
            game_name = td_evt[0].text.strip('\n')
            game_names.extend((game_name, game_name))
            # 开赛时间
            td_end_time = tr.find_all('td', class_="td td-endtime")
            game_time = td_end_time[0].text
            game_times.extend((game_time, game_time))
            # 依次为主队排名':'','主队':'','比分':'','客队':'','客队排名': ''
            td_team = tr.find_all('td', class_="td td-team")
            i = td_team[0].find_all('i')
            i0.extend((i[0].text, ''))
            i1.extend((i[2].text, ''))
            a = td_team[0].find_all('a')
            if len(a) == 3:
                a0.extend((a[0].text, ''))
                a1.extend((a[1].text, ''))
                a2.extend((a[2].text, ''))
            # 比赛未进行标签数量有变
            if len(a) < 3:
                a0.extend((a[0].text, ''))
                a1.extend((i[1].text, ''))
                a2.extend((a[1].text, ''))
            # 让球
            td_rang = tr.find_all('td', class_="td td-rang")
            p = td_rang[0].find_all('p')
            p0 = p[0].text
            p1 = p[1].text
            concede.extend((p0, p1))
            # 赔率
            td_bet_btn = tr.find_all('td', class_="td td-betbtn")
            p = td_bet_btn[0].find_all('p')
            if len(p) == 6:
                # 同时加载多个元素到列表
                victory.extend((p[0].text, p[3].text))
                # 加载字体标识判断值到列表
                victory_index.extend((p[0].attrs['class'][-1], p[3].attrs['class'][-1]))
                flat.extend((p[1].text, p[4].text))
                flat_index.extend((p[1].attrs['class'][-1], p[4].attrs['class'][-1]))
                defeat.extend((p[2].text, p[5].text))
                defeat_index.extend((p[2].attrs['class'][-1], p[5].attrs['class'][-1]))
            else:
                # 同时加载多个元素到列表
                victory.extend(('未开售', p[0].text))
                victory_index.extend(('未开售', p[0].attrs['class'][-1]))
                flat.extend(('未开售', p[1].text))
                flat_index.extend(('未开售', p[1].attrs['class'][-1]))
                defeat.extend(('未开售', p[2].text))
                defeat_index.extend(('未开售', p[2].attrs['class'][-1]))

    data['胜'], data['平'], data['负'] = victory, flat, defeat
    title['编号'], title['赛事'], title['开赛时间'], title['让球'] = serial_numbers, game_names, game_times, concede
    title['主队排名'], title['主队'], title['比分'], title['客队'], title['客队排名'] = i0, a0, a1, a2, i1
    p_victory = pd.Series(data=victory, index=victory_index)
    p_flat = pd.Series(data=flat, index=flat_index)
    p_defeat = pd.Series(data=defeat, index=defeat_index)
    df1 = pd.DataFrame(title)
    df2 = pd.DataFrame(data)

    now = time.time()
    ls = time.localtime(now)
    # filetime = time.strftime('%Y_%m_%d')
    filetime = x_date
    # df = pd.merge(df2,df1,how="inner",left_index=True,right_index=True)
    df1.to_excel('{}.xlsx'.format(filetime))
    # 通过p[n].attrs['class'] = ['betbtn'] or ['betbtn', 'betbtn-ok']确定中奖赔率(如何带格式写入)
    # 打开Excel程序，默认设置：程序可见，只打开不新建工作薄，屏幕更新关闭
    app = xw.App(visible=True, add_book=False)
    app.screen_updating = False  # 关闭显示更新
    app.display_alerts = False  # 关闭提示信息
    wk = app.books.open('{}.xlsx'.format(filetime))
    sht1 = wk.sheets['sheet1']
    sht1.range('K1').value = '胜'
    sht1.range('L1').value = '平'
    sht1.range('M1').value = '负'
    sht1.range('K1:M1').api.Font.Bold = True
    for nrows in range(1, len(p_flat) + 1):
        sht1.range(nrows + 1, 11).value = p_victory[nrows - 1]
        sht1.range(nrows + 1, 12).value = p_flat[nrows - 1]
        sht1.range(nrows + 1, 13).value = p_defeat[nrows - 1]
        if p_victory.index[nrows - 1] == 'betbtn-ok':
            sht1.range(nrows + 1, 11).api.Font.ColorIndex = 3
            sht1.range(nrows + 1, 11).api.Font.Bold = True
        if p_flat.index[nrows - 1] == 'betbtn-ok':
            sht1.range(nrows + 1, 12).api.Font.ColorIndex = 3
            sht1.range(nrows + 1, 12).api.Font.Bold = True
        if p_defeat.index[nrows - 1] == 'betbtn-ok':
            sht1.range(nrows + 1, 13).api.Font.ColorIndex = 3
            sht1.range(nrows + 1, 13).api.Font.Bold = True
    # 居中
    sht1.range('A1:M{}'.format(len(p_flat) + 1)).api.HorizontalAlignment = -4108
    sht1.range('A1:M{}'.format(len(p_flat) + 1)).api.VerticalAlignment = -4108
    # 增加边框
    sht1.range('A1:M{}'.format(len(p_flat) + 1)).api.Borders(9).LineStyle = 1
    sht1.range('A1:M{}'.format(len(p_flat) + 1)).api.Borders(10).LineStyle = 1
    sht1.range('A1:M{}'.format(len(p_flat) + 1)).api.Borders(11).LineStyle = 1
    sht1.range('A1:M{}'.format(len(p_flat) + 1)).api.Borders(12).LineStyle = 1
    sht1.autofit()
    # 合并单元格
    n_clos = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
    nrows = sht1.used_range.last_cell.row
    for nclo in n_clos:
        for nrow in range(2, nrows, 2):
            cell = nclo + str(nrow) + ':' + nclo + str((nrow + 1))
            sht1.range(cell).merge()
    wk.save()
    wk.close()
    app.quit()



# def analysis_daily_data(x_date):
#     """
#         分析每天的比赛变化
#     :return:
#     """
#     num = 8
#     for i in range(num):
#         if i == 0:
#             print('waiting for analysis')
#         if i == num - 1:
#             print('🥇')
#         else:
#             print('⚽', end='')
#         time.sleep(1)
#     daily_data_file = x_date + '.xlsx'
#
#     workbook = openpyxl.load_workbook(daily_data_file)
#     # active_sheet = workbook.active
#     sheets = workbook.sheetnames
#     worksheet = workbook[sheets[0]]  # 读取第一个sheet表格
#     worksheet.title = '处理数据表'  # 修改表名
#
#     print(f'sheet的最大行数: {worksheet.max_row}')
#     print(f'sheet的最大列数: {worksheet.max_column}')
#     copied_sheet = workbook.copy_worksheet(worksheet)
#     copied_sheet.title = "原始数据表"  # 备份原始数据表
#
#     num_zheng = 0
#     num_fan = 0
#     num_ping = 0
#
#     game_zheng_league = []
#     game_fan_league = []
#     game_ping_league = []
#
#     # 判断是否是正路
#     for num in range(2, worksheet.max_row, 2):
#         game_cell = 'C' + str(num)
#         score_cell = 'G' + str(num)
#         win_cell = 'K' + str(num)
#         ping_cell = 'L' + str(num)
#         lose_cell = 'M' + str(num)
#         res_win_cell = 'K' + str(num + 1)
#         res_ping_cell = 'L' + str(num + 1)
#         res_lose_cell = 'M' + str(num + 1)
#
#         score = worksheet[score_cell].value
#         if score == 'VS':
#             continue
#         else:
#             zhu_score = int(score.split(':')[0])
#             ke_score = int(score.split(':')[1])
#
#         if worksheet[win_cell].value != '未开售' and worksheet[ping_cell].value != '未开售' and worksheet[lose_cell].value != '未开售':
#             win_sp = float(worksheet[win_cell].value)
#             ping_sp = float(worksheet[ping_cell].value)
#             lose_sp = float(worksheet[lose_cell].value)
#             minimum = min(win_sp, ping_sp, lose_sp)  # 获取胜平负三者赔率的最小值
#         else:
#             continue
#
#         # 合并单元格
#         start_cell = 'N' + str(num)
#         end_cell = 'N' + str(num + 1)
#         range_string = start_cell + ':' + end_cell
#         worksheet.merge_cells(range_string)
#         color = ['ffc7ce', '4785f4', 'ffa500']  # 粉红、天蓝、橘黄
#
#         if zhu_score > ke_score:
#             if minimum == win_sp:
#                 game_zheng_league.append(worksheet[game_cell].value)
#
#                 num_zheng += 1
#                 worksheet[start_cell] = '正路'
#                 # 填充背景色
#                 fill_cell = PatternFill('solid', fgColor=color[0])
#                 worksheet[start_cell].fill = fill_cell
#
#             if minimum == ping_sp:
#                 worksheet[start_cell] = '反路-平'
#                 # 填充背景色
#                 fill_cell = PatternFill('solid', fgColor=color[1])
#                 worksheet[start_cell].fill = fill_cell
#
#             if minimum == lose_sp:
#                 game_fan_league.append(worksheet[game_cell].value)
#
#                 num_fan += 1
#                 worksheet[start_cell] = '反路'
#                 # 填充背景色
#                 fill_cell = PatternFill('solid', fgColor=color[2])
#                 worksheet[start_cell].fill = fill_cell
#
#         elif zhu_score < ke_score:
#             if minimum == win_sp:
#                 game_fan_league.append(worksheet[game_cell].value)
#
#                 num_fan += 1
#                 worksheet[start_cell] = '反路'
#
#                 # 填充背景色
#                 fill_cell = PatternFill('solid', fgColor=color[2])
#                 worksheet[start_cell].fill = fill_cell
#
#             if minimum == ping_sp:
#                 worksheet[start_cell] = '反路-平'
#
#                 # 填充背景色
#                 fill_cell = PatternFill('solid', fgColor=color[1])
#                 worksheet[start_cell].fill = fill_cell
#
#             if minimum == lose_sp:
#                 game_zheng_league.append(worksheet[game_cell].value)
#
#                 num_zheng += 1
#                 worksheet[start_cell] = '正路'
#
#                 # 填充背景色
#                 fill_cell = PatternFill('solid', fgColor=color[0])
#                 worksheet[start_cell].fill = fill_cell
#
#         else:
#             game_ping_league.append(worksheet[game_cell].value)
#
#             worksheet[start_cell] = '平'
#             num_ping += 1
#
#             opt_start_cell = 'O' + str(num)
#             worksheet[opt_start_cell] = '反路'
#             opt_end_cell = 'O' + str(num + 1)
#             r_string = opt_start_cell + ':' + opt_end_cell
#             worksheet.merge_cells(r_string)
#             fill_cell = PatternFill('solid', fgColor='e4e861')
#             worksheet[opt_start_cell].fill = fill_cell
#
#             # 填充背景色
#             fill_cell = PatternFill('solid', fgColor='18eb0c')
#             worksheet[start_cell].fill = fill_cell
#
#     # print('正路赛事: ', game_zheng_league)
#     # print('反路赛事: ', game_fan_league)
#     # print('平局赛事: ', game_ping_league)
#
#     print('正路:', num_zheng)
#     print('反路:', num_fan)
#     print('平局:', num_ping)
#     print('共计比赛数:', int(int(worksheet.max_row) - 1) / 2)
#     summary_txt = '正路: ' + str(num_zheng) + '\n' + '反路: ' + str(num_fan) + '\n' + '平局: ' + str(num_ping) + '\n' + '比赛数: ' + str(int(int(worksheet.max_row) - 1) / 2)
#     worksheet['Q2'].alignment = Alignment(wrapText=True)
#     worksheet['Q2'] = summary_txt
#     fill_Q2 = PatternFill('solid', fgColor='7dd1ec')
#     worksheet['Q2'].fill = fill_Q2
#     font = Font(u'微软雅黑', size=13, bold=True, italic=False, strike=False)
#     worksheet['Q2'].font = font
#
#     range_string = 'Q2' + ':' + 'S17'
#     worksheet.merge_cells(range_string)
#
#     # 居中单元格
#     alignment_center = Alignment(horizontal='center', vertical='center')
#     column_alp = chr(ord('A') + worksheet.max_column - 1)
#     range_end_cell = column_alp + str(worksheet.max_row)
#     range_cell = 'A1:' + range_end_cell
#     # print(f'全表单元格范围: {range_cell}')
#     for row_cell in worksheet[range_cell]:
#         for cell in row_cell:
#             cell.alignment = alignment_center
#
#     new_xlsx_file = 'proceed_' + x_date + '.xlsx'
#     workbook.save(new_xlsx_file)
#     workbook.close()
#     print(f'表格保存路径: {new_xlsx_file}')
#
#     # 调用show_it 函数
#     for _ in range(6):
#         print('😊', end='')
#         time.sleep(0.5)
#     show_it(game_zheng_league, game_fan_league, game_ping_league, 'proceed_' + x_date)
#
#     file_path = x_date + ".xlsx"  # 将 "your_excel_file.xlsx" 替换为你要删除的Excel文件的路径
#     if os.path.exists(file_path):
#         os.remove(file_path)
#         print(f"{file_path} 已成功删除。")
#     else:
#         print(f"{file_path} 不存在，无法删除。")


def main():
    x_date = '2024-04-12'
    search_func(x_date)
    # analysis_daily_data(x_date)


if __name__ == '__main__':
    main()
