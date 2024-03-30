"""
    从Full_data.xlsx中获取单关数据并处理后用柱状图表示出来,
    生成 single_match_game.xlsx excel文件 和 single_match.png 柱状图
        有用
        直接运行
"""

import os
import time
import numpy as np
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt


def single_show_it(positive_sports, negative_sports, draw_sports, x_date):
    # 设置Matplotlib的字体为支持中文字符的字体，例如'SimHei'或'Microsoft YaHei'
    plt.rcParams['font.sans-serif'] = ['SimHei']

    # # 正路赛事数据
    # positive_sports = ['日乙', '意甲', '西甲', '俄超', '意甲', '西甲', '挪超', '意甲', '葡超', '美职足']
    # # 反路赛事数据
    # negative_sports = ['英超', '瑞超', '挪超']
    # # 平局赛事数据
    # draw_sports = ['法甲', '德乙', '德甲', '英超', '瑞超', '西甲']

    # 定义固定的联赛顺序
    # fixed_order = ['俄超', '英超', '瑞超', '挪超', '法甲', '德乙', '德甲', '意甲', '日乙', '西甲', '葡超', '美职足']
    all_sports = positive_sports + negative_sports + draw_sports
    fixed_order = list(set(all_sports))

    # 初始化正路赛事、反路赛事和平局赛事的数量列表
    positive_counts = []
    negative_counts = []
    draw_counts = []
    # 计算每个联赛的正路赛事、反路赛事和平局赛事数量
    for sport in fixed_order:
        positive_count = positive_sports.count(sport)
        negative_count = negative_sports.count(sport)
        draw_count = draw_sports.count(sport)

        if positive_count != 0 or negative_count != 0 or draw_count != 0:
            positive_counts.append(positive_count)
            negative_counts.append(negative_count)
            draw_counts.append(draw_count)

    # 设置柱状图的宽度
    bar_width = 0.25

    # 设置x轴位置
    x = np.arange(len(fixed_order))

    # 创建一个新的画布，并指定画布的大小
    plt.figure(figsize=(12, 6))

    # 创建柱状图
    plt.bar(x, positive_counts, width=bar_width, label='正路赛事')
    plt.bar(x + bar_width, negative_counts, width=bar_width, label='反路赛事')
    plt.bar(x + 2 * bar_width, draw_counts, width=bar_width, label='平局赛事')

    # 设置x轴标签
    plt.xticks(x + bar_width, fixed_order, rotation=90)
    # plt.gcf().autofmt_xdate()

    # 设置图例
    plt.legend()

    # 设置图表标题和坐标轴标签
    plt.title('正路赛事、反路赛事和平局赛事统计')
    plt.xlabel('联赛名称')
    plt.ylabel('出现次数')

    # 设置纵坐标刻度间隔为1
    plt.yticks(range(0, max(positive_counts + negative_counts + draw_counts) + 1, 1))

    # 在柱子顶部显示数量
    for i, count in enumerate(positive_counts):
        if count != 0:
            plt.text(x[i], count, str(count), ha='center', va='bottom')
    for i, count in enumerate(negative_counts):
        if count != 0:
            plt.text(x[i] + bar_width, count, str(count), ha='center', va='bottom')
    for i, count in enumerate(draw_counts):
        if count != 0:
            plt.text(x[i] + 2 * bar_width, count, str(count), ha='center', va='bottom')

    # 保存柱状图为图片
    png_name = x_date + '.png'
    plt.savefig(png_name, dpi=300, bbox_inches='tight')

    # 显示柱状图
    plt.tight_layout()
    plt.show()


def analysis_single_data(single_data_file):
    """
        分析单关的数据变化
    :return:
    """
    num = 8
    for i in range(num):
        if i == 0:
            print('waiting for analysis')
        if i == num - 1:
            print('🥇')
        else:
            print('⚽', end='')
        time.sleep(1)

    workbook = openpyxl.load_workbook(single_data_file)
    # active_sheet = workbook.active
    sheets = workbook.sheetnames
    worksheet = workbook[sheets[0]]  # 读取第一个sheet表格
    worksheet.title = '处理数据表'  # 修改表名

    print(f'sheet的最大行数: {worksheet.max_row}')
    print(f'sheet的最大列数: {worksheet.max_column}')
    # copied_sheet = workbook.copy_worksheet(worksheet)
    # copied_sheet.title = "原始数据表"  # 备份原始数据表

    num_zheng = 0
    num_fan = 0
    num_ping = 0

    game_zheng_league = []
    game_fan_league = []
    game_ping_league = []

    # 判断是否是正路
    for num in range(1, worksheet.max_row, 2):
        game_cell = 'C' + str(num)
        score_cell = 'G' + str(num)
        win_cell = 'K' + str(num)
        ping_cell = 'L' + str(num)
        lose_cell = 'M' + str(num)
        res_win_cell = 'K' + str(num + 1)
        res_ping_cell = 'L' + str(num + 1)
        res_lose_cell = 'M' + str(num + 1)

        score = worksheet[score_cell].value
        if score == 'VS':
            continue
        else:
            zhu_score = int(score.split(':')[0])
            ke_score = int(score.split(':')[1])

        if worksheet[win_cell].value != '未开售' and worksheet[ping_cell].value != '未开售' and worksheet[
            lose_cell].value != '未开售':
            win_sp = float(worksheet[win_cell].value)
            ping_sp = float(worksheet[ping_cell].value)
            lose_sp = float(worksheet[lose_cell].value)
            minimum = min(win_sp, ping_sp, lose_sp)  # 获取胜平负三者赔率的最小值
        else:
            continue

        # 合并单元格
        start_cell = 'N' + str(num)
        end_cell = 'N' + str(num + 1)
        range_string = start_cell + ':' + end_cell
        worksheet.merge_cells(range_string)
        color = ['ffc7ce', '4785f4', 'ffa500']  # 粉红、天蓝、橘黄

        if zhu_score > ke_score:
            if minimum == win_sp:
                game_zheng_league.append(worksheet[game_cell].value)

                num_zheng += 1
                worksheet[start_cell] = '正路'
                # 填充背景色
                fill_cell = PatternFill('solid', fgColor=color[0])
                worksheet[start_cell].fill = fill_cell

            if minimum == ping_sp:
                worksheet[start_cell] = '反路-平'
                # 填充背景色
                fill_cell = PatternFill('solid', fgColor=color[1])
                worksheet[start_cell].fill = fill_cell

            if minimum == lose_sp:
                game_fan_league.append(worksheet[game_cell].value)

                num_fan += 1
                worksheet[start_cell] = '反路'
                # 填充背景色
                fill_cell = PatternFill('solid', fgColor=color[2])
                worksheet[start_cell].fill = fill_cell

        elif zhu_score < ke_score:
            if minimum == win_sp:
                game_fan_league.append(worksheet[game_cell].value)

                num_fan += 1
                worksheet[start_cell] = '反路'

                # 填充背景色
                fill_cell = PatternFill('solid', fgColor=color[2])
                worksheet[start_cell].fill = fill_cell

            if minimum == ping_sp:
                worksheet[start_cell] = '反路-平'

                # 填充背景色
                fill_cell = PatternFill('solid', fgColor=color[1])
                worksheet[start_cell].fill = fill_cell

            if minimum == lose_sp:
                game_zheng_league.append(worksheet[game_cell].value)

                num_zheng += 1
                worksheet[start_cell] = '正路'

                # 填充背景色
                fill_cell = PatternFill('solid', fgColor=color[0])
                worksheet[start_cell].fill = fill_cell

        else:
            game_ping_league.append(worksheet[game_cell].value)

            worksheet[start_cell] = '平'
            num_ping += 1

            opt_start_cell = 'O' + str(num)
            worksheet[opt_start_cell] = '反路'
            opt_end_cell = 'O' + str(num + 1)
            r_string = opt_start_cell + ':' + opt_end_cell
            worksheet.merge_cells(r_string)
            fill_cell = PatternFill('solid', fgColor='e4e861')
            worksheet[opt_start_cell].fill = fill_cell

            # 填充背景色
            fill_cell = PatternFill('solid', fgColor='18eb0c')
            worksheet[start_cell].fill = fill_cell

    # print('正路赛事: ', game_zheng_league)
    # print('反路赛事: ', game_fan_league)
    # print('平局赛事: ', game_ping_league)

    print('正路:', num_zheng)
    print('反路:', num_fan)
    print('平局:', num_ping)
    print('共计比赛数:', int(int(worksheet.max_row) / 2))
    summary_txt = '正路: ' + str(num_zheng) + '\n' + '反路: ' + str(num_fan) + '\n' + '平局: ' + str(
        num_ping) + '\n' + '比赛数: ' + str(int(int(worksheet.max_row) / 2))
    worksheet['Q2'].alignment = Alignment(wrapText=True)
    worksheet['Q2'] = summary_txt
    fill_Q2 = PatternFill('solid', fgColor='7dd1ec')
    worksheet['Q2'].fill = fill_Q2
    font = Font(u'微软雅黑', size=13, bold=True, italic=False, strike=False)
    worksheet['Q2'].font = font

    range_string = 'Q2' + ':' + 'S17'
    worksheet.merge_cells(range_string)

    # 居中单元格
    alignment_center = Alignment(horizontal='center', vertical='center')
    column_alp = chr(ord('A') + worksheet.max_column - 1)
    range_end_cell = column_alp + str(worksheet.max_row)
    range_cell = 'A1:' + range_end_cell
    # print(f'全表单元格范围: {range_cell}')
    for row_cell in worksheet[range_cell]:
        for cell in row_cell:
            cell.alignment = alignment_center

    new_xlsx_file = 'single_match_game' + '.xlsx'
    workbook.save(new_xlsx_file)
    workbook.close()
    print(f'表格保存路径: {new_xlsx_file}')

    # 调用show_it 函数
    for _ in range(6):
        print('😊', end='')
        time.sleep(0.5)
    single_show_it(game_zheng_league, game_fan_league, game_ping_league, 'single_match')

    file_path = "单关数据表.xlsx"  # 将 "your_excel_file.xlsx" 替换为你要删除的Excel文件的路径
    if os.path.exists(file_path):
        os.remove(file_path)
        print(f"{file_path} 已成功删除。")
    else:
        print(f"{file_path} 不存在，无法删除。")


def optim_xlsx(excel_file):
    """
        优化excel表格
        合并 居中 单元格
    :return:
    """
    for _ in range(6):
        print('😊', end='')
        time.sleep(0.5)

    # 合并 居中
    print('😊😊😊😊😊😊😊😊😊😊😊')
    workbook = openpyxl.load_workbook(excel_file)
    # active_sheet = workbook.active
    sheets = workbook.sheetnames
    worksheet = workbook[sheets[0]]  # 读取第一个sheet表格
    worksheet.title = '处理数据表'  # 修改表名

    # print(f'sheet的最大行数: {worksheet.max_row}')
    # print(f'sheet的最大列数: {worksheet.max_column}')
    copied_sheet = workbook.copy_worksheet(worksheet)
    copied_sheet.title = "原始数据表"  # 备份原始数据表

    # 合并单元格
    # columns_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']  # 需要合并单元格修改的列
    columns_list = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']  # 需要合并单元格修改的列
    for col in columns_list:
        for i in range(1, worksheet.max_row + 1, 2):
            start_cell = col + str(i)
            end_cell = col + str(i + 1)
            range_string = start_cell + ':' + end_cell
            # print(f'{range_string}')
            worksheet.merge_cells(range_string)

    # 标记
    font = Font(color='FF0000', bold=True)

    for num in range(1, worksheet.max_row, 2):
        score_cell = 'G' + str(num)
        res_win_cell = 'K' + str(num)
        res_ping_cell = 'L' + str(num)
        res_lose_cell = 'M' + str(num)
        # print(score_cell)
        score = worksheet[score_cell].value
        # print(score)
        if score == 'VS':
            continue
        zhu_score = int(score.split(':')[0])
        ke_score = int(score.split(':')[1])

        # 让球
        rang_score_cell = 'J' + str(num + 1)
        res_rang_win_cell = 'K' + str(num + 1)
        res_rang_ping_cell = 'L' + str(num + 1)
        res_rang_lose_cell = 'M' + str(num + 1)
        rang_score = str(worksheet[rang_score_cell].value)
        rang_score = int(rang_score)

        if zhu_score > ke_score:
            # print('win')
            # 标记胜平负
            worksheet[res_win_cell].font = font
            if zhu_score + rang_score > ke_score:
                worksheet[res_rang_win_cell].font = font  # 标记让胜
            elif zhu_score + rang_score == ke_score:
                worksheet[res_rang_ping_cell].font = font  # 标记让平
            else:
                worksheet[res_rang_lose_cell].font = font  # 标记让负

        elif zhu_score == ke_score:
            # print('ping')
            # 标记胜平负
            worksheet[res_ping_cell].font = font
            if zhu_score + rang_score > ke_score:
                worksheet[res_rang_win_cell].font = font  # 标记让胜
            elif zhu_score + rang_score == ke_score:
                worksheet[res_rang_ping_cell].font = font  # 标记让平
            else:
                worksheet[res_rang_lose_cell].font = font  # 标记让负

        else:
            # print('lose')
            # 标记胜平负
            worksheet[res_lose_cell].font = font
            if zhu_score + rang_score > ke_score:
                worksheet[res_rang_win_cell].font = font  # 标记让胜
            elif zhu_score + rang_score == ke_score:
                worksheet[res_rang_ping_cell].font = font  # 标记让平
            else:
                worksheet[res_rang_lose_cell].font = font  # 标记让负

    # 居中单元格
    alignment_center = Alignment(horizontal='center', vertical='center')
    column_alp = chr(ord('A') + worksheet.max_column - 1)
    range_end_cell = column_alp + str(worksheet.max_row)
    range_cell = 'A1:' + range_end_cell
    # print(f'全表单元格范围: {range_cell}')
    for row_cell in worksheet[range_cell]:
        for cell in row_cell:
            cell.alignment = alignment_center

    new_excel_file = '单关数据表.xlsx'
    workbook.save(new_excel_file)
    workbook.close()
    print(f'表格保存路径: {new_excel_file}')

    file_path = "Single_Match_Data.xlsx"  # 将 "your_excel_file.xlsx" 替换为你要删除的Excel文件的路径

    if os.path.exists(file_path):
        os.remove(file_path)
        print(f"{file_path} 已成功删除。")
    else:
        print(f"{file_path} 不存在，无法删除。")

    # 生成柱状图
    analysis_single_data(new_excel_file)


def get_single_match():
    """
        从 Full_data.xlsx 表中获取单关数据
    :return:
    """
    num = 3
    for i in range(num):
        if i == 0:
            print('waiting for analysis')
        if i == num - 1:
            print('🥇')
        else:
            print('⚽', end='')
        time.sleep(1)

    full_data_file = 'all_games.xlsx'
    workbook = openpyxl.load_workbook(full_data_file)
    # active_sheet = workbook.active
    sheets = workbook.sheetnames
    worksheet = workbook[sheets[0]]  # 读取第一个sheet表格
    worksheet.title = '处理数据表'  # 修改表名

    print(f'sheet的最大行数: {worksheet.max_row}')
    print(f'sheet的最大列数: {worksheet.max_column}')
    # copied_sheet = workbook.copy_worksheet(worksheet)
    # copied_sheet.title = "原始数据表"  # 备份原始数据表

    # 创建一个新的工作簿和工作表
    new_workbook = openpyxl.Workbook()
    new_worksheet = new_workbook.active
    new_worksheet.title = 'single_match_sheet'

    # 判断是否是单关
    for num in range(worksheet.max_row - 1, 1, -2):  # 逆序遍历行
        single_match_cell = 'J' + str(num)
        is_single = worksheet[single_match_cell].value
        if is_single.startswith('单关'):
            # 提取当前行和下一行的数据
            row_data = [worksheet.cell(row=num, column=col).value for col in range(1, worksheet.max_column + 1)]
            next_row_data = [worksheet.cell(row=num + 1, column=col).value for col in
                             range(1, worksheet.max_column + 1)]

            # 向新工作表中添加数据
            new_worksheet.append(row_data)
            new_worksheet.append(next_row_data)

    # 保存新工作簿
    new_file = 'Single_Match_Data.xlsx'
    new_workbook.save(new_file)
    new_workbook.close()

    # 美化excel表格
    optim_xlsx(new_file)


if __name__ == '__main__':
    get_single_match()
