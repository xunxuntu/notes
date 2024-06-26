# -*- coding: utf-8 -*-
"""
    @File : search_low_1.5.py
    @Time : 2024-02-26
    Description : 
"""

import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from tqdm import tqdm
from datetime import datetime
from pathlib import Path


def mark_handicap_pos_neg_outcome(xlsx_path, sheet_name, output_xlsx_file):
    """
        计算赛果的盘口胜平负(正反路)
    :return:
    """

    workbook = openpyxl.load_workbook(xlsx_path)
    worksheet = workbook[sheet_name]

    # print(worksheet.dimensions) # 获取表格的尺寸大小(几行几列数据)

    # color = ['ffa500', '4785f4']  # 橘黄(反路)、天蓝(正路)
    color = ['EEB4B4', '00FF7F']  # 红色和绿色的十六进制颜色值

    num_zheng = 0
    num_fan = 0
    num_rang_sheng = 0
    num_rang_ping = 0
    num_rang_fu = 0

    for num in range(2, worksheet.max_row, 2):
        # print(f'num: {num}')
        score_cell = 'G' + str(num)
        handicap_cell = 'J' + str(num + 1)
        score = worksheet[score_cell].value
        handicap = worksheet[handicap_cell].value

        handicap_win_cell = 'K' + str(num + 1)
        handicap_ping_cell = 'L' + str(num + 1)
        handicap_lose_cell = 'M' + str(num + 1)
        handicap_win_sp = worksheet[handicap_win_cell].value
        handicap_ping_sp = worksheet[handicap_ping_cell].value
        handicap_lose_sp = worksheet[handicap_lose_cell].value

        try:
            minimum = min(handicap_win_sp, handicap_ping_sp, handicap_lose_sp)
            maximum = max(handicap_win_sp, handicap_ping_sp, handicap_lose_sp)

            home_score = int(score.split(':')[0])  # 主场进球数
            away_score = int(score.split(':')[1])  # 客场进球数

            fill_cell = 'N' + str(num + 1)

            if home_score > away_score:
                # print(f'胜')
                if home_score + int(handicap) > away_score:
                    # print(f'让胜')
                    if minimum == handicap_win_sp:
                        num_zheng += 1
                        # print(f'num: {num}, 让胜, handicap 正路')
                        # 填充背景色
                        worksheet[fill_cell] = 'handicap 正路'
                        fg_color = PatternFill('solid', fgColor=color[1])
                        worksheet[fill_cell].fill = fg_color

                    elif maximum == handicap_win_sp:
                        num_fan += 1
                        num_rang_sheng += 1
                        # print(f'num: {num}, 让胜, handicap 反路')
                        # 填充背景色
                        worksheet[fill_cell] = 'handicap 反路-让胜'
                        fg_color = PatternFill('solid', fgColor=color[0])
                        worksheet[fill_cell].fill = fg_color

                elif home_score + int(handicap) == away_score:
                    # print(f'让平')
                    if minimum != handicap_ping_sp:
                        num_fan += 1
                        num_rang_ping += 1
                        # print(f'num: {num}, 让平, handicap 反路')
                        # 填充背景色
                        worksheet[fill_cell] = 'handicap 反路-让平'
                        fg_color = PatternFill('solid', fgColor=color[0])
                        worksheet[fill_cell].fill = fg_color

                else:
                    # print(f'让负')
                    if minimum == handicap_lose_sp:
                        num_zheng += 1
                        # print(f'num: {num}, 让负, handicap 正路')
                        # 填充背景色
                        worksheet[fill_cell] = 'handicap 正路'
                        fg_color = PatternFill('solid', fgColor=color[1])
                        worksheet[fill_cell].fill = fg_color

            elif home_score == away_score:
                # print(f'平')
                if home_score + int(handicap) > away_score:
                    # print(f'让胜')
                    if minimum == handicap_win_sp:
                        num_zheng += 1
                        # print(f'num: {num}, 平, handicap 正路')
                        # 填充背景色
                        worksheet[fill_cell] = 'handicap 正路'
                        fg_color = PatternFill('solid', fgColor=color[1])
                        worksheet[fill_cell].fill = fg_color

                elif home_score + int(handicap) == away_score:
                    # print(f'让平')
                    if minimum != handicap_ping_sp:
                        num_fan += 1
                        num_rang_ping += 1
                        # print(f'num: {num}, 让平, handicap 反路')
                        # 填充背景色
                        worksheet[fill_cell] = 'handicap 反路'
                        fg_color = PatternFill('solid', fgColor=color[0])
                        worksheet[fill_cell].fill = fg_color
                else:
                    # print(f'让负')
                    if minimum == handicap_lose_sp:
                        num_zheng += 1
                        # print(f'num: {num}, 让负, handicap 正路')
                        # 填充背景色
                        worksheet[fill_cell] = 'handicap 正路'
                        fg_color = PatternFill('solid', fgColor=color[1])
                        worksheet[fill_cell].fill = fg_color

            else:
                # print(f'负')
                if home_score + int(handicap) > away_score:
                    # print(f'让胜')
                    if minimum == handicap_win_sp:
                        num_zheng += 1
                        # print(f'num: {num}, 让胜, handicap 正路')
                        # 填充背景色
                        worksheet[fill_cell] = 'handicap 正路'
                        fg_color = PatternFill('solid', fgColor=color[1])
                        worksheet[fill_cell].fill = fg_color

                elif home_score + int(handicap) == away_score:
                    # print(f'让平')
                    if minimum != handicap_ping_sp:
                        num_fan += 1
                        num_rang_ping += 1
                        # print(f'num: {num}, 让平, handicap 反路')
                        # 填充背景色
                        worksheet[fill_cell] = 'handicap 反路-让平'
                        fg_color = PatternFill('solid', fgColor=color[0])
                        worksheet[fill_cell].fill = fg_color

                else:
                    # print(f'让负')
                    if minimum == handicap_lose_sp:
                        num_zheng += 1
                        # print(f'num: {num}, 让负, handicap 正路')
                        # 填充背景色
                        worksheet[fill_cell] = 'handicap 正路'
                        fg_color = PatternFill('solid', fgColor=color[1])
                        worksheet[fill_cell].fill = fg_color

                    if minimum == handicap_win_sp:
                        num_fan += 1
                        num_rang_fu += 1
                        # print(f'num: {num}, 让负, handicap 反路')
                        # 填充背景色
                        worksheet[fill_cell] = 'handicap 反路-让负'
                        fg_color = PatternFill('solid', fgColor=color[0])
                        worksheet[fill_cell].fill = fg_color

        except (TypeError, ValueError):
            # print(f'{num} 比分的格式错误，请检查该场比赛的比分字符串')
            pass

    print()
    print(f'handicap 正路： {num_zheng} 场, 占比 {round((num_zheng / (num_zheng + num_fan)) * 100, 1)}% ')
    print(f'handicap 反路： {num_fan}场, 占比 {round((num_fan / (num_zheng + num_fan)) * 100, 1)}% ')

    print(f'handicap 反路-让胜： {num_rang_sheng}')
    print(f'handicap 反路-让平： {num_rang_ping}')
    print(f'handicap 反路-让负： {num_rang_fu}')

    print('共计比赛数:', int(num_zheng + num_fan))
    # print('共计行数:', int(worksheet.max_row))

    # 保存新文件
    workbook.save(output_xlsx_file)
    print(f'表格保存路径: {output_xlsx_file}')


def mark_pos_neg_outcome(xlsx_path, sheet_name, output_xlsx_file):
    """
        标记赛果的胜平负
    :return:
    """
    workbook = openpyxl.load_workbook(xlsx_path)
    worksheet = workbook[sheet_name]

    # print(worksheet.dimensions) # 获取表格的尺寸大小(几行几列数据)

    color = ['ffc0cb', '7fff00']  # 粉红(反路)、嫩绿(正路)

    num_zheng = 0
    num_fan = 0

    for num in range(2, worksheet.max_row, 2):
        score_cell = 'G' + str(num)
        handicap_cell = 'J' + str(num + 1)
        score = worksheet[score_cell].value
        handicap = worksheet[handicap_cell].value

        win_cell = 'K' + str(num)
        ping_cell = 'L' + str(num)
        lose_cell = 'M' + str(num)
        win_sp = worksheet[win_cell].value
        ping_sp = worksheet[ping_cell].value
        lose_sp = worksheet[lose_cell].value

        try:
            if win_sp == '未开售':
                continue

            minimum = min(win_sp, ping_sp, lose_sp)

            home_score = int(score.split(':')[0])  # 主场进球数
            away_score = int(score.split(':')[1])  # 客场进球数

            fill_cell = 'N' + str(num)

            if home_score > away_score:
                if minimum == win_sp:
                    num_zheng += 1
                    # print(f'num: {num}, 胜, sp 正路')
                    worksheet[fill_cell] = '正路'
                    # 填充背景色
                    fg_color = PatternFill('solid', fgColor=color[1])
                    worksheet[fill_cell].fill = fg_color

                if minimum == ping_sp:
                    num_fan += 1
                    worksheet[fill_cell] = '反路-胜'
                    # 填充背景色
                    fg_color = PatternFill('solid', fgColor=color[0])
                    worksheet[fill_cell].fill = fg_color

                if minimum == lose_sp:
                    num_fan += 1
                    worksheet[fill_cell] = '反路-胜'
                    # 填充背景色
                    fg_color = PatternFill('solid', fgColor=color[0])
                    worksheet[fill_cell].fill = fg_color

            elif home_score < away_score:
                if minimum == win_sp:
                    num_fan += 1
                    worksheet[fill_cell] = '反路-负'

                    # 填充背景色
                    fg_color = PatternFill('solid', fgColor=color[0])
                    worksheet[fill_cell].fill = fg_color

                if minimum == ping_sp:
                    num_fan += 1
                    worksheet[fill_cell] = '反路-平'

                    # 填充背景色
                    fg_color = PatternFill('solid', fgColor=color[0])
                    worksheet[fill_cell].fill = fg_color

                if minimum == lose_sp:
                    num_zheng += 1
                    worksheet[fill_cell] = '正路'

                    # 填充背景色
                    fg_color = PatternFill('solid', fgColor=color[1])
                    worksheet[fill_cell].fill = fg_color

            else:
                if minimum == ping_sp:
                    num_zheng += 1
                    worksheet[fill_cell] = '正路'
                    # 填充背景色
                    fg_color = PatternFill('solid', fgColor=color[1])
                    worksheet[fill_cell].fill = fg_color

                if minimum == lose_sp:
                    num_fan += 1
                    worksheet[fill_cell] = '反路-平'
                    # 填充背景色
                    fg_color = PatternFill('solid', fgColor=color[0])
                    worksheet[fill_cell].fill = fg_color

                if minimum == win_sp:
                    num_fan += 1
                    worksheet[fill_cell] = '反路-平'
                    # 填充背景色
                    fg_color = PatternFill('solid', fgColor=color[0])
                    worksheet[fill_cell].fill = fg_color

        except (TypeError, ValueError):
            # print(f'比分的格式错误，请检查该场比赛的比分字符串')
            pass

    print()
    print(f'正路： {num_zheng} 场, 占比 {round((num_zheng / (num_zheng + num_fan)) * 100, 2)}%')
    print(f'反路： {num_fan} 场, 占比 {round((num_fan / (num_zheng + num_fan)) * 100, 2)}%')

    print('共计比赛数:', int(num_zheng + num_fan))
    # print('共计行数:', int(worksheet.max_row))

    # 保存新文件
    workbook.save(output_xlsx_file)


def search_specify_handicap_sp(xlsx_path, sheet_name, output_xlsx_file):
    """
        在 all_games。xlsx 文件中查找自定义范围 handicap_sp 值的场次
    :return:
    """
    workbook = openpyxl.load_workbook(xlsx_path)
    worksheet = workbook[sheet_name]

    # print(worksheet.dimensions) # 获取表格的尺寸大小(几行几列数据)

    # 获取受让胜平负的sp
    need_del_row = []
    for num in range(2, worksheet.max_row, 2):
        # print(num)
        handicap_win_cell = 'K' + str(num + 1)
        handicap_ping_cell = 'L' + str(num + 1)
        handicap_lose_cell = 'M' + str(num + 1)
        handicap_win_sp = worksheet[handicap_win_cell].value
        handicap_ping_sp = worksheet[handicap_ping_cell].value
        handicap_lose_sp = worksheet[handicap_lose_cell].value
        # print(handicap_win_sp, handicap_ping_sp, handicap_lose_sp)

        minimum_handicap_sp = min(handicap_win_sp, handicap_ping_sp, handicap_lose_sp)

        # ## 保留 handicap_sp 小于1.5的场次
        # 先找出 handicap_sp 大于1.5的场次，再删除所有大于1.5的场次，剩下的就是小于1.5的场次了
        if minimum_handicap_sp > 1.4:
            need_del_row.append(num)
            need_del_row.append(num + 1)

        # if minimum_handicap_sp < 1.4 or minimum_handicap_sp > 1.5:
        #     # print(f'num {num} 大于1.5， {handicap_win_sp, handicap_ping_sp, handicap_lose_sp}')
        #     need_del_row.append(num)
        #     need_del_row.append(num + 1)

    # 从后往前删除
    # print(need_del_row)
    print(f'正在删除不符合条件的行 ... ')
    for i in tqdm(need_del_row[::-1]):
        worksheet.delete_rows(i)  # 表示删除表格的第i行

    # 保存新文件
    workbook.save(output_xlsx_file)

    """
    # 获取比分
    score = worksheet['G3208'].value
    print(score)

    # 获取受让球数
    handicap = worksheet['J3209'].value
    print(handicap)

    # 获取胜平负和受让胜平负的sp
    win_sp = worksheet['K3208'].value
    draw_sp = worksheet['L3208'].value
    loss_sp = worksheet['M3208'].value
    handicap_win_score = worksheet['K3209'].value
    handicap_draw_score = worksheet['L3209'].value
    handicap_loss_score = worksheet['M3209'].value
    match_sp = [win_sp, draw_sp, loss_sp, handicap_win_score, handicap_draw_score, handicap_loss_score]
    print(*match_sp)
    """


def search_specify_sp(xlsx_path, sheet_name, output_xlsx_file):
    """
        # 在 all_games.xlsx 查找自定义范围 sp 值的场次
    :return:
    """
    workbook = openpyxl.load_workbook(xlsx_path)
    worksheet = workbook[sheet_name]

    # print(worksheet.dimensions) # 获取表格的尺寸大小(几行几列数据)

    # 获取胜平负的sp
    need_del_row = []
    for num in range(2, worksheet.max_row, 2):
        # print(num)
        win_cell = 'K' + str(num)
        ping_cell = 'L' + str(num)
        lose_cell = 'M' + str(num)
        win_sp = worksheet[win_cell].value
        ping_sp = worksheet[ping_cell].value
        lose_sp = worksheet[lose_cell].value
        # print(win_sp, ping_sp, lose_sp)

        # 过滤"未开售"
        if win_sp == '未开售':
            continue

        minimum_sp = min(win_sp, ping_sp, lose_sp)

        # ## 保留 sp 小于1.5的场次
        # 先找出 sp 大于1.5的场次，再删除所有大于1.5的场次，剩下的就是小于1.5的场次了
        if minimum_sp > 1.5:
            # print(f'num {num} 大于1.5, {win_sp, ping_sp, lose_sp}')
            need_del_row.append(num)
            need_del_row.append(num + 1)

    # 从后往前删除
    # print(need_del_row)
    for i in need_del_row[::-1]:
        print(f'正在删除第 {i} 行 ... ')
        worksheet.delete_rows(i)  # 表示删除表格的第i行

    # 保存新文件
    workbook.save(output_xlsx_file)


def count_non_empty_rows(file_path, sheet_name):
    """
        计算excel中的非空白行的数量
    :param file_path:
    :param sheet_name:
    :return:
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]

    non_empty_row_count = 0

    # 遍历表格中的每一行
    for row in range(1, ws.max_row + 1):
        # 检查每一行中是否至少有一个单元格非空
        if any([cell.value is not None for cell in ws[row]]):
            non_empty_row_count += 1

    return non_empty_row_count


def format_excel(excel_file):
    """
        美化 excel
    :return:
    """
    workbook = openpyxl.load_workbook(excel_file)
    sheets = workbook.sheetnames
    worksheet = workbook[sheets[0]]  # 读取第一个sheet表格

    # print(f'sheet的最大行数: {worksheet.max_row}')
    # print(f'sheet的最大列数: {worksheet.max_column}')

    # 合并单元格
    columns_list = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']  # 需要合并单元格修改的列
    for col in columns_list:
        for i in range(2, worksheet.max_row + 1, 2):
            start_cell = col + str(i)
            end_cell = col + str(i + 1)
            range_string = start_cell + ':' + end_cell
            # print(f'{range_string}')
            worksheet.merge_cells(range_string)

    # 标记赛果 sp
    font = Font(color='FF0000', bold=True)
    for num in range(2, worksheet.max_row, 2):
        score_cell = 'G' + str(num)
        win_cell = 'K' + str(num)
        ping_cell = 'L' + str(num)
        lose_cell = 'M' + str(num)
        score = worksheet[score_cell].value
        if score == 'VS':
            continue
        home_score = int(score.split(':')[0])
        away_score = int(score.split(':')[1])

        # 让球
        handicap_score_cell = 'J' + str(num + 1)
        handicap_win_cell = 'K' + str(num + 1)
        handicap_ping_cell = 'L' + str(num + 1)
        handicap_lose_cell = 'M' + str(num + 1)
        handicap_score = str(worksheet[handicap_score_cell].value)
        if len(handicap_score) > 4:  # 过滤出 " 单关-3" 这样的字符
            if handicap_score[:3] == " 单关":
                handicap_score = int(handicap_score[3:])
        else:
            handicap_score = int(handicap_score)

        if home_score > away_score:
            # 标记胜平负
            if worksheet[win_cell].value == '未开售':
                pass
            else:
                worksheet[win_cell].font = font
            if home_score + handicap_score > away_score:
                worksheet[handicap_win_cell].font = font  # 标记让胜
            elif home_score + handicap_score == away_score:
                worksheet[handicap_ping_cell].font = font  # 标记让平
            else:
                worksheet[handicap_lose_cell].font = font  # 标记让负

        elif home_score == away_score:
            # 标记胜平负
            worksheet[ping_cell].font = font
            if home_score + handicap_score > away_score:
                worksheet[handicap_win_cell].font = font  # 标记让胜
            elif home_score + handicap_score == away_score:
                worksheet[handicap_ping_cell].font = font  # 标记让平
            else:
                worksheet[handicap_lose_cell].font = font  # 标记让负

        else:
            # 标记胜平负
            worksheet[lose_cell].font = font
            if home_score + handicap_score > away_score:
                worksheet[handicap_win_cell].font = font  # 标记让胜
            elif home_score + handicap_score == away_score:
                worksheet[handicap_ping_cell].font = font  # 标记让平
            else:
                worksheet[handicap_lose_cell].font = font  # 标记让负

    # 居中单元格, 绘制边框线
    # 创建边框线对象
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
    # 居中单元格
    alignment_center = Alignment(horizontal='center', vertical='center')
    column_alp = chr(ord('A') + worksheet.max_column - 1)
    range_end_cell = column_alp + str(worksheet.max_row)
    range_cell = 'A1:' + range_end_cell
    # print(f'全表单元格范围: {range_cell}')
    for row_cell in worksheet[range_cell]:
        for cell in row_cell:
            cell.alignment = alignment_center
            cell.border = border

    # 保存表格
    workbook.save(excel_file)
    workbook.close()
    print(f'美化表格保存路径: {excel_file}')


def find_custom_sp_matches(custom_sp_excel,
                           win_sp=None, win_sp_fluctuation=None,
                           draw_sp=None, draw_sp_fluctuation=None,
                           lose_sp=None, lose_sp_fluctuation=None,
                           handicap_win_sp=None, handicap_win_sp_fluctuation=None,
                           handicap_draw_sp=None, handicap_draw_sp_fluctuation=None,
                           handicap_lose_sp=None, handicap_lose_sp_fluctuation=None
                           ):
    """
        在 all_games.xlsx 中查找相近 sp 值的场次
        【使用pandas实现】
    :return:
    """
    xlsx_file = r'all_games.xlsx'
    # xlsx_file = r'澳超0412.xlsx'
    # df = pd.read_excel(xlsx_file, sheet_name='all')
    df = pd.read_excel(xlsx_file, sheet_name='all')

    # 数据格式转换
    df['胜'] = pd.to_numeric(df['胜'], errors='coerce')
    df['负'] = pd.to_numeric(df['负'], errors='coerce')
    df['平'] = pd.to_numeric(df['平'], errors='coerce')

    # 查找 让胜 和 让负 的 sp 值
    # handicap_win_sp = df.loc[df['序号'] % 2 != 0, ['序号', '胜']]
    # handicap_lose_sp = df.loc[df['序号'] % 2 != 0, ['序号', '负']]
    # print(handicap_win_sp)
    # print(handicap_lose_sp)

    # 筛选自定义 sp 值的相似值的场次
    custom_sp_df = pd.DataFrame()
    # print(f'excel 表格有 3245 行，DataFrame 有 {len(df)} 行')  # DataFrame的行数计算是不包括列名这一行的。

    for i in range(0, len(df)):
        if i % 2 == 0:  # 筛选相同 胜 sp
            # 筛选 win_sp
            if win_sp:
                if draw_sp:  # 胜 平
                    if win_sp_fluctuation and draw_sp_fluctuation:
                        if win_sp - win_sp_fluctuation <= df.loc[i, '胜'] <= win_sp + win_sp_fluctuation \
                                and draw_sp - draw_sp_fluctuation <= df.loc[i, '平'] <= draw_sp + draw_sp_fluctuation:
                            custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

                    elif win_sp_fluctuation:
                        if win_sp - win_sp_fluctuation <= df.loc[i, '胜'] <= win_sp + win_sp_fluctuation \
                                and draw_sp == df.loc[i, '平']:
                            custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

                    elif draw_sp_fluctuation:
                        if win_sp == df.loc[i, '胜'] \
                                and draw_sp - draw_sp_fluctuation <= df.loc[i, '平'] <= draw_sp + draw_sp_fluctuation:
                            custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

                    elif win_sp == df.loc[i, '胜'] and draw_sp == df.loc[i, '平']:
                        custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

                elif lose_sp:  # 胜 负
                    if win_sp_fluctuation and lose_sp_fluctuation:
                        if win_sp - win_sp_fluctuation <= df.loc[i, '胜'] <= win_sp + win_sp_fluctuation \
                                and lose_sp - lose_sp_fluctuation <= df.loc[i, '负'] <= lose_sp + lose_sp_fluctuation:
                            custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

                    elif win_sp_fluctuation:
                        if win_sp - win_sp_fluctuation <= df.loc[i, '胜'] <= win_sp + win_sp_fluctuation \
                                and lose_sp == df.loc[i, '负']:
                            custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

                    elif lose_sp_fluctuation:
                        if win_sp == df.loc[i, '胜'] \
                                and lose_sp - lose_sp_fluctuation <= df.loc[i, '负'] <= lose_sp + lose_sp_fluctuation:
                            custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

                    elif win_sp == df.loc[i, '胜'] and lose_sp == df.loc[i, '负']:
                        custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

                elif handicap_lose_sp:  # 胜 让负
                    if win_sp_fluctuation and handicap_lose_sp_fluctuation:
                        if win_sp - win_sp_fluctuation <= df.loc[i, '胜'] <= win_sp + win_sp_fluctuation \
                                and handicap_lose_sp - handicap_lose_sp_fluctuation <= df.loc[i + 1, '负'] <= \
                                handicap_lose_sp + handicap_lose_sp_fluctuation:
                            custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

                    elif win_sp_fluctuation:
                        if win_sp - win_sp_fluctuation <= df.loc[i, '胜'] <= win_sp + win_sp_fluctuation \
                                and handicap_lose_sp == df.loc[i + 1, '负']:
                            custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

                    elif handicap_lose_sp_fluctuation:
                        if win_sp == df.loc[i, '胜'] and handicap_lose_sp - handicap_lose_sp_fluctuation <= \
                                df.loc[i + 1, '负'] <= handicap_lose_sp + handicap_lose_sp_fluctuation:
                            custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

                    elif win_sp == df.loc[i, '胜'] and handicap_lose_sp == df.loc[i + 1, '负']:
                        custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

                elif win_sp_fluctuation:
                    if win_sp - win_sp_fluctuation <= df.loc[i, '胜'] <= win_sp + win_sp_fluctuation:
                        custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)
                else:
                    if win_sp == df.loc[i, '胜']:
                        custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

            # 筛选 draw_sp
            elif draw_sp:
                if lose_sp:  # 平 负
                    if draw_sp_fluctuation and lose_sp_fluctuation:
                        if draw_sp - draw_sp_fluctuation <= df.loc[i, '平'] <= draw_sp + draw_sp_fluctuation \
                                and lose_sp - lose_sp_fluctuation <= df.loc[i, '负'] <= lose_sp + lose_sp_fluctuation:
                            custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

                    elif draw_sp_fluctuation:
                        if draw_sp - draw_sp_fluctuation <= df.loc[i, '平'] <= draw_sp + draw_sp_fluctuation \
                                and win_sp == df.loc[i, '胜']:
                            custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

                    elif lose_sp_fluctuation:
                        if draw_sp == df.loc[i, '平'] \
                                and lose_sp - lose_sp_fluctuation <= df.loc[i, '负'] <= lose_sp + lose_sp_fluctuation:
                            custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

                    elif win_sp == df.loc[i, '胜'] and lose_sp == df.loc[i, '负']:
                        custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

                elif handicap_draw_sp:  # 平 让平
                    if draw_sp_fluctuation and handicap_draw_sp_fluctuation:
                        if draw_sp - draw_sp_fluctuation <= df.loc[i, '平'] <= draw_sp + draw_sp_fluctuation \
                                and handicap_draw_sp - handicap_draw_sp_fluctuation <= \
                                df.loc[i + 1, '平'] <= handicap_draw_sp + handicap_draw_sp_fluctuation:
                            custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

                    elif draw_sp_fluctuation:
                        if draw_sp - draw_sp_fluctuation <= df.loc[i, '平'] <= draw_sp + draw_sp_fluctuation \
                                and handicap_draw_sp == df.loc[i + 1, '平']:
                            custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

                    elif handicap_draw_sp_fluctuation:
                        if draw_sp == df.loc[i, '平'] and handicap_draw_sp - handicap_draw_sp_fluctuation <= \
                                df.loc[i + 1, '平'] <= handicap_draw_sp + handicap_draw_sp_fluctuation:
                            custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

                    elif draw_sp == df.loc[i, '平'] and handicap_draw_sp == df.loc[i + 1, '平']:
                        custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

                elif draw_sp_fluctuation:
                    if draw_sp - draw_sp_fluctuation <= df.loc[i, '平'] <= draw_sp + draw_sp_fluctuation:
                        custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

                else:
                    if draw_sp == df.loc[i, '平']:
                        custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

            # 筛选 lose_sp
            elif lose_sp:
                if handicap_win_sp:  # 负 让胜
                    if lose_sp_fluctuation and handicap_win_sp_fluctuation:
                        if lose_sp - lose_sp_fluctuation <= df.loc[i, '负'] <= lose_sp + lose_sp_fluctuation \
                                and handicap_win_sp - handicap_win_sp_fluctuation <= \
                                df.loc[i + 1, '胜'] <= handicap_win_sp + handicap_win_sp_fluctuation:
                            custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

                    elif lose_sp_fluctuation:
                        if lose_sp - lose_sp_fluctuation <= df.loc[i, '负'] <= lose_sp + lose_sp_fluctuation \
                                and handicap_win_sp == df.loc[i + 1, '胜']:
                            custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

                    elif handicap_win_sp_fluctuation:
                        if handicap_draw_sp - handicap_draw_sp_fluctuation <= df.loc[i + 1, '胜'] <= \
                                handicap_draw_sp + handicap_draw_sp_fluctuation and lose_sp == df.loc[i, '负']:
                            custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

                    elif lose_sp == df.loc[i, '负'] and handicap_win_sp == df.loc[i + 1, '胜']:
                        custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

                elif lose_sp_fluctuation:
                    if lose_sp - lose_sp_fluctuation <= df.loc[i, '负'] <= lose_sp + lose_sp_fluctuation:
                        custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

                else:
                    if lose_sp == df.loc[i, '负']:
                        custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

            # 筛选 handicap_win_sp
            elif handicap_win_sp:
                if handicap_win_sp_fluctuation:
                    if handicap_win_sp - handicap_win_sp_fluctuation <= df.loc[i + 1, '胜'] <= \
                            handicap_win_sp + handicap_win_sp_fluctuation:
                        custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)
                else:
                    if handicap_win_sp == df.loc[i + 1, '胜']:
                        custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

            elif handicap_draw_sp:
                if handicap_draw_sp_fluctuation:
                    if handicap_draw_sp - handicap_draw_sp_fluctuation <= df.loc[i + 1, '平'] <= \
                            handicap_draw_sp + handicap_draw_sp_fluctuation:
                        custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)
                else:
                    if handicap_draw_sp == df.loc[i + 1, '平']:
                        custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

            elif handicap_lose_sp:
                if handicap_lose_sp_fluctuation:
                    if handicap_lose_sp - handicap_lose_sp_fluctuation <= df.loc[i + 1, '负'] <= \
                            handicap_lose_sp + handicap_lose_sp_fluctuation:
                        custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)
                else:
                    if handicap_lose_sp == df.loc[i + 1, '负']:
                        custom_sp_df = custom_sp_df.append([df.loc[i], df.loc[i + 1]], ignore_index=True)

    # 导出为 excel
    # custom_sp_excel = 'custom_sp_df.xlsx'
    custom_sp_df.to_excel(custom_sp_excel, index=False)

    # 美化 excel
    format_excel(custom_sp_excel)

    # sp 正反路标记统计
    mark_pos_neg_outcome(xlsx_path=custom_sp_excel,
                         sheet_name="Sheet1",
                         output_xlsx_file=custom_sp_excel)

    print(f'表格保存路径: {custom_sp_excel}')


def find_special_sp_matches(threshold_num):
    """
        查找特殊 sp 值的比赛
    :return:
    """
    xlsx_file = r'all_games.xlsx'
    df = pd.read_excel(xlsx_file, sheet_name='all')

    # 筛选 handicap_sp 值低于1.5的场次
    # 查找 让胜 或者 让负 在低于 1.5 之下的df
    # threshold_num = 1.4
    special_sp_df = pd.DataFrame()
    for i in range(1, len(df), 2):
        # print(i)
        # print(df.loc[i, '负'])
        # print(type(df.loc[i, '负']))
        if df.loc[i, '负'] <= threshold_num or df.loc[i, '胜'] <= threshold_num:
            special_sp_df = special_sp_df.append([df.loc[i - 1], df.loc[i]], ignore_index=True)

    # 导出为 excel
    current_date = datetime.now().date()
    current_date_path = Path(str(current_date))
    current_date_path.mkdir(exist_ok=True)
    special_sp_excel = current_date_path / f'受让球sp小于{threshold_num}场次-{current_date}.xlsx'
    special_sp_df.to_excel(special_sp_excel, index=False)

    # 美化 excel
    format_excel(special_sp_excel)

    # handicap sp 正反路标记统计
    mark_handicap_pos_neg_outcome(xlsx_path=special_sp_excel,
                                  sheet_name="Sheet1",
                                  output_xlsx_file=special_sp_excel)


def find_around_sp_matches():
    """
        查找胜 or 负 sp在1.4以下的比赛
    :return:
    """
    xlsx_file = r'all_games.xlsx'
    df = pd.read_excel(xlsx_file, sheet_name='all')

    # 筛选 胜or负 sp 值低于1.4的场次
    threshold_num = 1.4
    special_sp_df = pd.DataFrame()
    for i in range(1, len(df), 2):
        # print(i)
        # print(df.loc[i-1, '负'])
        # print(type(df.loc[i, '负']))
        if df.loc[i - 1, '负'] == '未开售':
            continue
        if df.loc[i - 1, '负'] <= threshold_num or df.loc[i - 1, '胜'] <= threshold_num:
            special_sp_df = special_sp_df.append([df.loc[i - 1], df.loc[i]], ignore_index=True)

    # 导出为 excel
    special_sp_excel = '胜or负sp小于1.4场次.xlsx'
    special_sp_df.to_excel(special_sp_excel, index=False)

    # 美化 excel
    format_excel(special_sp_excel)

    # sp 正反路标记统计
    mark_pos_neg_outcome(xlsx_path=special_sp_excel,
                         sheet_name="Sheet1",
                         output_xlsx_file=special_sp_excel)


# def match_analysis(win_sp=None, win_sp_fluctuation=None,
#                    draw_sp=None, draw_sp_fluctuation=None,
#                    lose_sp=None, lose_sp_fluctuation=None,
#                    handicap_win_sp=1.54, handicap_win_sp_fluctuation=None,
#                    handicap_draw_sp=None, handicap_draw_sp_fluctuation=None,
#                    handicap_lose_sp=None, handicap_lose_sp_fluctuation=None):
#     """
#         场次分析
#     :return:
#     """


def get_double_draw(excel_file=r'all_games.xlsx'):
    """
    获取比赛中双平的场次
    :return:
    """
    # excel_file = r'all_games.xlsx'
    current_date = datetime.now().date()
    current_date_path = Path(str(current_date))
    current_date_path.mkdir(exist_ok=True)
    output_xlsx_file = current_date_path / f"{excel_file.split(':')[0]}-双平场次-{current_date}.xlsx"

    # 读取原始 Excel 文件
    df = pd.read_excel(excel_file)

    # 新建一个 DataFrame 用于存储符合条件的数据
    new_df = pd.DataFrame(columns=df.columns)

    num_range = range(0, len(df), 2)  # 适用于每行两个球队比赛的情况
    for num in tqdm(num_range, desc='Processing'):  # 使用 tqdm 显示处理进度
        # print(f'num: {num}')
        score = df.at[num, '比分']
        # print(f'score: {score}')
        if pd.isna(score):
            continue  # 跳过NaN值的处理
        if score == 'VS':
            continue
        home_score, away_score = map(int, str(score).split(':'))
        # print(f'home_score: {home_score}, away_score: {away_score}')

        # 让球
        handicap_score = df.at[num + 1, '让球']
        if isinstance(handicap_score, str) and len(handicap_score) > 4:  # 过滤出 " 单关-3" 这样的字符
            if handicap_score[:3] == " 单关":
                handicap_score = int(handicap_score[3:])
        else:
            handicap_score = int(handicap_score)

        if home_score + handicap_score == away_score or home_score == away_score:
            # 双平局
            new_df = new_df.append(df.iloc[num:num + 2])

    match_count = new_df['比分'].count()
    all_match_count = df['比分'].count()
    percentage = match_count / all_match_count * 100
    print(f"双平场次: {match_count}, 占比: {percentage:.1f}%", )

    # 保存新文件
    new_df.to_excel(output_xlsx_file, index=False)

    # 美化 excel
    format_excel(output_xlsx_file)


def get_loss_handicap_loss():
    """
        获取比赛中负 让负的比赛场次
    :return:
    """
    excel_file = r'all_games.xlsx'
    output_xlsx_file = 'pd-furangfu.xlsx'

    # 读取原始 Excel 文件
    df = pd.read_excel(excel_file)

    # 新建一个 DataFrame 用于存储符合条件的数据
    new_df = pd.DataFrame(columns=df.columns)

    num_range = range(0, len(df), 2)  # 适用于每行两个球队比赛的情况
    for num in tqdm(num_range, desc='Processing'):  # 使用 tqdm 显示处理进度
        print(f'num: {num}')
        score = df.at[num, '比分']
        print(f'score: {score}')
        if pd.isna(score):
            continue  # 跳过NaN值的处理
        if score == 'VS':
            continue
        home_score, away_score = map(int, str(score).split(':'))
        print(f'home_score: {home_score}, away_score: {away_score}')

        # 让球
        handicap_score = df.at[num + 1, '让球']
        if isinstance(handicap_score, str) and len(handicap_score) > 4:  # 过滤出 " 单关-3" 这样的字符
            if handicap_score[:3] == " 单关":
                handicap_score = int(handicap_score[3:])
        else:
            handicap_score = int(handicap_score)

        if home_score < away_score and home_score + handicap_score < away_score:
            # 负 让负局
            new_df = new_df.append(df.iloc[num:num + 2])

    # 保存新文件
    new_df.to_excel(output_xlsx_file, index=False)

    # 美化 excel
    format_excel(output_xlsx_file)


def get_win_handicap_win():
    """
       获取比赛中胜-让胜的比赛场次
    :return:
    """
    excel_file = r'all_games.xlsx'
    output_xlsx_file = 'pd-shrangsh2.xlsx'

    # 读取原始 Excel 文件
    df = pd.read_excel(excel_file)

    # 新建一个 DataFrame 用于存储符合条件的数据
    new_df = pd.DataFrame(columns=df.columns)

    num_range = range(0, len(df), 2)  # 适用于每行两个球队比赛的情况
    for num in tqdm(num_range, desc='Processing'):  # 使用 tqdm 显示处理进度
        print(f'num: {num}')
        score = df.at[num, '比分']
        print(f'score: {score}')
        if pd.isna(score):
            continue  # 跳过NaN值的处理
        if score == 'VS':
            continue
        home_score, away_score = map(int, str(score).split(':'))
        print(f'home_score: {home_score}, away_score: {away_score}')

        # 让球
        handicap_score = df.at[num + 1, '让球']
        if isinstance(handicap_score, str) and len(handicap_score) > 4:  # 过滤出 " 单关-3" 这样的字符
            if handicap_score[:3] == " 单关":
                handicap_score = int(handicap_score[3:])
        else:
            handicap_score = int(handicap_score)

        if home_score > away_score and home_score + handicap_score > away_score:
            # 胜 让胜局
            new_df = new_df.append(df.iloc[num:num + 2])

    # 保存新文件
    new_df.to_excel(output_xlsx_file, index=False)

    # 美化 excel
    format_excel(output_xlsx_file)


def get_win_handicap_lose():
    """
        获取比赛中胜-让负的比赛场次
    :return:
    """
    excel_file = r'all_games.xlsx'
    output_xlsx_file = 'pd-shrangfu.xlsx'

    # 读取原始 Excel 文件
    df = pd.read_excel(excel_file)

    # 新建一个 DataFrame 用于存储符合条件的数据
    new_df = pd.DataFrame(columns=df.columns)

    num_range = range(0, len(df), 2)  # 适用于每行两个球队比赛的情况
    for num in tqdm(num_range, desc='Processing'):  # 使用 tqdm 显示处理进度
        print(f'num: {num}')
        score = df.at[num, '比分']
        print(f'score: {score}')
        if pd.isna(score):
            continue  # 跳过NaN值的处理
        if score == 'VS':
            continue
        home_score, away_score = map(int, str(score).split(':'))
        print(f'home_score: {home_score}, away_score: {away_score}')

        # 让球
        handicap_score = df.at[num + 1, '让球']
        if isinstance(handicap_score, str) and len(handicap_score) > 4:  # 过滤出 " 单关-3" 这样的字符
            if handicap_score[:3] == " 单关":
                handicap_score = int(handicap_score[3:])
        else:
            handicap_score = int(handicap_score)

        if home_score > away_score and home_score + handicap_score < away_score:
            # 胜 让负局
            new_df = new_df.append(df.iloc[num:num + 2])

    # 保存新文件
    new_df.to_excel(output_xlsx_file, index=False)

    # 美化 excel
    format_excel(output_xlsx_file)


def get_lose_handicap_win():
    """
        获取比赛中负-让胜的比赛场次
    :return:
    """
    excel_file = r'all_games.xlsx'
    output_xlsx_file = 'pd-furangsh.xlsx'

    # 读取原始 Excel 文件
    df = pd.read_excel(excel_file)

    # 新建一个 DataFrame 用于存储符合条件的数据
    new_df = pd.DataFrame(columns=df.columns)

    num_range = range(0, len(df), 2)  # 适用于每行两个球队比赛的情况
    for num in tqdm(num_range, desc='Processing'):  # 使用 tqdm 显示处理进度
        print(f'num: {num}')
        score = df.at[num, '比分']
        print(f'score: {score}')
        if pd.isna(score):
            continue  # 跳过NaN值的处理
        if score == 'VS':
            continue
        home_score, away_score = map(int, str(score).split(':'))
        print(f'home_score: {home_score}, away_score: {away_score}')

        # 让球
        handicap_score = df.at[num + 1, '让球']
        if isinstance(handicap_score, str) and len(handicap_score) > 4:  # 过滤出 " 单关-3" 这样的字符
            if handicap_score[:3] == " 单关":
                handicap_score = int(handicap_score[3:])
        else:
            handicap_score = int(handicap_score)

        if home_score < away_score and home_score + handicap_score > away_score:
            # 负 让胜局
            new_df = new_df.append(df.iloc[num:num + 2])

    # 保存新文件
    new_df.to_excel(output_xlsx_file, index=False)

    # 美化 excel
    format_excel(output_xlsx_file)


def get_handicap_lose():
    """
        获取比赛中让负的比赛场次
    :return:
    """
    excel_file = r'all_games.xlsx'
    output_xlsx_file = 'pd-rangfu.xlsx'

    # 读取原始 Excel 文件
    df = pd.read_excel(excel_file)

    # 新建一个 DataFrame 用于存储符合条件的数据
    new_df = pd.DataFrame(columns=df.columns)

    num_range = range(0, len(df), 2)  # 适用于每行两个球队比赛的情况
    for num in tqdm(num_range, desc='Processing'):  # 使用 tqdm 显示处理进度
        print(f'num: {num}')
        score = df.at[num, '比分']
        print(f'score: {score}')
        if pd.isna(score):
            continue  # 跳过NaN值的处理
        if score == 'VS':
            continue
        home_score, away_score = map(int, str(score).split(':'))
        print(f'home_score: {home_score}, away_score: {away_score}')

        # 让球
        handicap_score = df.at[num + 1, '让球']
        if isinstance(handicap_score, str) and len(handicap_score) > 4:  # 过滤出 " 单关-3" 这样的字符
            if handicap_score[:3] == " 单关":
                handicap_score = int(handicap_score[3:])
        else:
            handicap_score = int(handicap_score)

        if home_score + handicap_score < away_score:
            # 让负局
            new_df = new_df.append(df.iloc[num:num + 2])

    # 保存新文件
    new_df.to_excel(output_xlsx_file, index=False)

    # 美化 excel
    format_excel(output_xlsx_file)


def get_handicap_win():
    """
        获取比赛中让胜的比赛场次
    :return:
    """
    excel_file = r'all_games.xlsx'
    output_xlsx_file = 'pd-rangsh.xlsx'

    # 读取原始 Excel 文件
    df = pd.read_excel(excel_file)

    # 新建一个 DataFrame 用于存储符合条件的数据
    new_df = pd.DataFrame(columns=df.columns)

    num_range = range(0, len(df), 2)  # 适用于每行两个球队比赛的情况
    for num in tqdm(num_range, desc='Processing'):  # 使用 tqdm 显示处理进度
        print(f'num: {num}')
        score = df.at[num, '比分']
        print(f'score: {score}')
        if pd.isna(score):
            continue  # 跳过NaN值的处理
        if score == 'VS':
            continue
        home_score, away_score = map(int, str(score).split(':'))
        print(f'home_score: {home_score}, away_score: {away_score}')

        # 让球
        handicap_score = df.at[num + 1, '让球']
        if isinstance(handicap_score, str) and len(handicap_score) > 4:  # 过滤出 " 单关-3" 这样的字符
            if handicap_score[:3] == " 单关":
                handicap_score = int(handicap_score[3:])
        else:
            handicap_score = int(handicap_score)

        if home_score + handicap_score > away_score:
            # 让胜局
            new_df = new_df.append(df.iloc[num:num + 2])

    # 保存新文件
    new_df.to_excel(output_xlsx_file, index=False)

    # 美化 excel
    format_excel(output_xlsx_file)


def get_handicap_draw():
    """
        获取比赛中让平的比赛场次
    :return:
    """
    excel_file = r'all_games.xlsx'
    output_xlsx_file = 'pd-rangdraw.xlsx'

    # 读取原始 Excel 文件
    df = pd.read_excel(excel_file)

    # 新建一个 DataFrame 用于存储符合条件的数据
    new_df = pd.DataFrame(columns=df.columns)

    num_range = range(0, len(df), 2)  # 适用于每行两个球队比赛的情况
    for num in tqdm(num_range, desc='Processing'):  # 使用 tqdm 显示处理进度
        print(f'num: {num}')
        score = df.at[num, '比分']
        print(f'score: {score}')
        if pd.isna(score):
            continue  # 跳过NaN值的处理
        if score == 'VS':
            continue
        home_score, away_score = map(int, str(score).split(':'))
        print(f'home_score: {home_score}, away_score: {away_score}')

        # 让球
        handicap_score = df.at[num + 1, '让球']
        if isinstance(handicap_score, str) and len(handicap_score) > 4:  # 过滤出 " 单关-3" 这样的字符
            if handicap_score[:3] == " 单关":
                handicap_score = int(handicap_score[3:])
        else:
            handicap_score = int(handicap_score)

        if home_score + handicap_score == away_score:
            # 让平局
            new_df = new_df.append(df.iloc[num:num + 2])

    # 保存新文件
    new_df.to_excel(output_xlsx_file, index=False)

    # 美化 excel
    format_excel(output_xlsx_file)


def find_rangp_2_ball_matches():
    """
        获取比赛中让平或者2球的比赛场次
    :return:
    """
    excel_file = r'all_games.xlsx'
    output_xlsx_file = 'pd-rangdraw-6balls.xlsx'

    # 读取原始 Excel 文件
    df = pd.read_excel(excel_file)

    # 新建一个 DataFrame 用于存储符合条件的数据
    new_df = pd.DataFrame(columns=df.columns)

    num_range = range(0, len(df), 2)  # 适用于每行两个球队比赛的情况
    for num in tqdm(num_range, desc='Processing'):  # 使用 tqdm 显示处理进度
        print(f'num: {num}')
        score = df.at[num, '比分']
        print(f'score: {score}')
        if pd.isna(score):
            continue  # 跳过NaN值的处理
        if score == 'VS':
            continue
        home_score, away_score = map(int, str(score).split(':'))
        print(f'home_score: {home_score}, away_score: {away_score}')

        # 让球
        handicap_score = df.at[num + 1, '让球']
        if isinstance(handicap_score, str) and len(handicap_score) > 4:  # 过滤出 " 单关-3" 这样的字符
            if handicap_score[:3] == " 单关":
                handicap_score = int(handicap_score[3:])
        else:
            handicap_score = int(handicap_score)

        if home_score + handicap_score == away_score or home_score + away_score == 6:
            # 让平 or 2球 的局
            new_df = new_df.append(df.iloc[num:num + 2])

    # 保存新文件
    new_df.to_excel(output_xlsx_file, index=False)

    # 美化 excel
    format_excel(output_xlsx_file)


def get_goat_matches():
    """
        获取进球数的分布
    :return:
    """
    excel_file = r'all_games.xlsx'
    current_date = datetime.now().date()
    current_date_path = Path(str(current_date))
    current_date_path.mkdir(exist_ok=True)
    output_xlsx_file = current_date_path / f'2-3球场次-{current_date}.xlsx'

    # 读取原始 Excel 文件
    df = pd.read_excel(excel_file)

    # 新建一个 DataFrame 用于存储符合条件的数据
    new_df = pd.DataFrame(columns=df.columns)
    num_range = range(0, len(df), 2)  # 适用于每行两个球队比赛的情况
    for num in tqdm(num_range, desc='Processing'):  # 使用 tqdm 显示处理进度
        # print(f'num: {num}')
        score = df.at[num, '比分']
        # print(f'score: {score}')
        if pd.isna(score):
            continue  # 跳过NaN值的处理
        if score == 'VS':
            continue

        home_score, away_score = map(int, str(score).split(':'))
        if home_score + away_score == 2 or home_score + away_score == 3:
            # 进球数是 2/3 球的局
            new_df = new_df.append(df.iloc[num:num + 2])

    # 保存新文件
    new_df.to_excel(output_xlsx_file, index=False)

    # 美化 excel
    format_excel(output_xlsx_file)


def auto_deal():
    """
        每日自动处理
        直接运行
    :return:
    """
    # 获取双平的场次
    get_double_draw()
    # 受让球sp小于1.4的场次
    find_special_sp_matches(1.4)
    # 受让球sp小于1.5的场次
    find_special_sp_matches(1.5)
    # 获取2-3球比赛场次
    get_goat_matches()


def find_handicap_two_goats():
    """
        # 获取所有让2球的比赛
    :return:
    """
    excel_file = r'all_games.xlsx'
    output_xlsx_file = r'handicap_two_balls.xlsx'
    output_xlsx_file_rang_ping = r'handicap_two_balls-让平局.xlsx'
    output_xlsx_file_rang_fu = r'handicap_two_balls-让负局.xlsx'
    output_xlsx_file_rang_sheng = r'handicap_two_balls-让胜局.xlsx'

    # 读取原始 Excel 文件
    df = pd.read_excel(excel_file)
    # 新建一个 DataFrame 用于存储符合条件的数据
    all_df = pd.DataFrame(columns=df.columns)
    rang_ping_df = pd.DataFrame(columns=df.columns)
    rang_fu_df = pd.DataFrame(columns=df.columns)
    rang_sheng_df = pd.DataFrame(columns=df.columns)

    num_range = range(0, len(df), 2)  # 适用于每行两个球队比赛的情况
    for num in tqdm(num_range, desc='Processing'):  # 使用 tqdm 显示处理进度
        handicap_ball = df.at[num + 1, '让球']
        score = df.at[num, '比分']

        if handicap_ball == ' -2' or handicap_ball == ' +2':
            all_df = all_df.append(df.iloc[num:num + 2])
            if pd.isna(score):
                continue  # 跳过NaN值的处理
            if score == 'VS':
                continue
            home_score, away_score = map(int, str(score).split(':'))

            # 让球
            # handicap_score = df.at[num + 1, '让球']
            # if isinstance(handicap_score, str) and len(handicap_score) > 4:  # 过滤出 " 单关-3" 这样的字符
            #     if handicap_score[:3] == " 单关":
            #         handicap_score = int(handicap_score[3:])
            # else:
            #     handicap_score = int(handicap_score)

            # if home_score + int(handicap_ball) == away_score:
            #     # 让平局
            #     rang_ping_df = rang_ping_df.append(df.iloc[num:num + 2])

            # if home_score + int(handicap_ball) < away_score:
            #     # 让负局
            #     rang_fu_df = rang_fu_df.append(df.iloc[num:num + 2])
            # #
            if home_score + int(handicap_ball) > away_score:
                # 让胜局
                rang_sheng_df = rang_sheng_df.append(df.iloc[num:num + 2])

    # 保存新文件
    all_df.to_excel(output_xlsx_file, index=False)
    # rang_ping_df.to_excel(output_xlsx_file_rang_ping, index=False)
    # rang_fu_df.to_excel(output_xlsx_file_rang_fu, index=False)
    rang_sheng_df.to_excel(output_xlsx_file_rang_sheng, index=False)

    # # 美化 excel
    format_excel(output_xlsx_file)
    # format_excel(output_xlsx_file_rang_ping)
    # format_excel(output_xlsx_file_rang_fu)
    format_excel(output_xlsx_file_rang_sheng)


def find_float_sp_matches(low_threshold_num, high_threshold_num):
    """
        查找区间内受让 sp 值的比赛
    :return:
    """
    xlsx_file = r'all_games.xlsx'
    df = pd.read_excel(xlsx_file, sheet_name='all')

    # 筛选 handicap_sp 值低于1.5的场次
    # 查找 让胜 或者 让负 在低于 1.5 之下的df
    # threshold_num = 1.4
    special_sp_df = pd.DataFrame()
    for i in range(1, len(df), 2):
        # print(i)
        # print(df.loc[i, '负'])
        # print(type(df.loc[i, '负']))
        if low_threshold_num <= df.loc[i, '负'] <= high_threshold_num or\
                low_threshold_num <= df.loc[i, '胜'] <= high_threshold_num:
            special_sp_df = special_sp_df.append([df.loc[i - 1], df.loc[i]], ignore_index=True)

    # 导出为 excel
    current_date = datetime.now().date()
    current_date_path = Path(str(current_date))
    current_date_path.mkdir(exist_ok=True)
    special_sp_excel = current_date_path / f'受让球sp大于{low_threshold_num}-小于{high_threshold_num} 场次-{current_date}.xlsx'
    special_sp_df.to_excel(special_sp_excel, index=False)

    # 美化 excel
    format_excel(special_sp_excel)

    # handicap sp 正反路标记统计
    mark_handicap_pos_neg_outcome(xlsx_path=special_sp_excel,
                                  sheet_name="Sheet1",
                                  output_xlsx_file=special_sp_excel)


if __name__ == '__main__':
    # 获取所有让2球的比赛
    # find_handicap_two_goats()

    # 场次分析
    # match_analysis(win_sp=2.2, win_sp_fluctuation=None,
    #                draw_sp=None, draw_sp_fluctuation=None,
    #                lose_sp=None, lose_sp_fluctuation=None,
    #                handicap_win_sp=None, handicap_win_sp_fluctuation=None,
    #                handicap_draw_sp=None, handicap_draw_sp_fluctuation=None,
    #                handicap_lose_sp=1.48, handicap_lose_sp_fluctuation=None)

    # 筛选自定义范围 sp 值的场次
    # outputs_excel = r'002.xlsx'
    # find_custom_sp_matches(outputs_excel,
    #                        win_sp=1.79, win_sp_fluctuation=0.02,
    #                        draw_sp=None, draw_sp_fluctuation=None,
    #                        lose_sp=None, lose_sp_fluctuation=None,
    #                        handicap_win_sp=None, handicap_win_sp_fluctuation=None,
    #                        handicap_draw_sp=None, handicap_draw_sp_fluctuation=None,
    #                        handicap_lose_sp=1.74, handicap_lose_sp_fluctuation=0.02)

    # 筛选特殊 sp 值的场次
    # 查找受让球sp小于1.4的场次
    # find_special_sp_matches(1.4)

    # 查找受让球sp小于1.5的场次
    # find_special_sp_matches(1.5)

    # 查找受让球sp小于1.5，大于1.4的场次
    # find_float_sp_matches(1.4, 1.5)

    # # # 查找自定义范围 sp 值的场次
    # search_specify_sp(xlsx_path=r'./all_games.xlsx', sheet_name="all",
    #                   output_xlsx_file='sp_around_1.5_game.xlsx')

    # # # 查找自定义范围 handicap_sp 值的场次
    # search_specify_handicap_sp(xlsx_path=r'./all_games.xlsx', sheet_name="all",
    #                            output_xlsx_file='让球低于1.4_game.xlsx')

    # 标记赛果的胜平负(正反路)
    # mark_pos_neg_outcome(xlsx_path='sp_around_1.5_game.xlsx',
    #                      sheet_name="all",
    #                      output_xlsx_file='sp_around_1.5_marked_pos_neg.xlsx')

    # # 标记赛果的盘口胜平负(正反路)
    # mark_handicap_pos_neg_outcome(xlsx_path='all_games.xlsx',
    #                               sheet_name="all",
    #                               output_xlsx_file='all_games_marked_handicap_pos_neg.xlsx')

    # 获取双平的比赛场次
    # get_double_draw(r'受让球sp大于1.4-小于1.5 场次-2024-04-24.xlsx')

    # 获取比赛中负-让负的比赛场次
    # get_loss_handicap_loss()

    # 获取比赛中胜-让胜的比赛场次
    # get_win_handicap_win()

    # 获取比赛中胜-让负的比赛场次
    # get_win_handicap_lose()

    # 获取比赛中负-让胜的比赛场次
    # get_lose_handicap_win()

    # 获取比赛中让负的比赛场次
    # get_handicap_lose()

    # 获取比赛中让胜的比赛场次
    # get_handicap_win()

    # 获取比赛中让平的比赛场次
    # get_handicap_draw()

    # 获取胜or负sp小于1.4的比赛场次
    find_around_sp_matches()

    # 获取让平或者2球的比赛场次
    # find_rangp_2_ball_matches()

    # 获取进球数的分布
    # get_goat_matches()

    # auto_deal()
