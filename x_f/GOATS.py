"""
有用
"""
import time
import requests
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from tqdm import tqdm


def format_excel(excel_file):
    """
        美化excel
    :return:
    """
    workbook = openpyxl.load_workbook(excel_file)
    sheets = workbook.sheetnames
    worksheet = workbook[sheets[0]]  # 读取第一个sheet表格

    # 合并单元格
    columns_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']  # 需要合并单元格修改的列
    for col in columns_list:
        for i in range(2, worksheet.max_row + 1, 2):
            start_cell = col + str(i)
            end_cell = col + str(i + 1)
            range_string = start_cell + ':' + end_cell
            # print(f'{range_string}')
            worksheet.merge_cells(range_string)

    # 标记赛果 sp
    font = Font(color='FF0000', bold=True)
    for num in range(2, worksheet.max_row + 1, 2):
        score_cell = 'C' + str(num)
        goat_0 = 'E' + str(num)
        goat_1 = 'F' + str(num)
        goat_2 = 'G' + str(num)
        goat_3 = 'H' + str(num)
        goat_4 = 'I' + str(num)
        goat_5 = 'J' + str(num)
        goat_6 = 'K' + str(num)
        goat_7 = 'L' + str(num)

        print(score_cell)
        score = worksheet[score_cell].value
        if score == 'VS':
            continue
        home_score = int(score.split(':')[0])
        away_score = int(score.split(':')[1])

        if home_score + away_score == 0:
            worksheet[goat_0].font = font
        elif home_score + away_score == 1:
            worksheet[goat_1].font = font
        elif home_score + away_score == 2:
            worksheet[goat_2].font = font
        elif home_score + away_score == 3:
            worksheet[goat_3].font = font
        elif home_score + away_score == 4:
            worksheet[goat_4].font = font
        elif home_score + away_score == 5:
            worksheet[goat_5].font = font
        elif home_score + away_score == 6:
            worksheet[goat_6].font = font
        elif home_score + away_score >= 7:
            worksheet[goat_7].font = font

    # 居中单元格, 绘制边框线
    # 创建边框线对象
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
    # 居中单元格
    alignment_center = Alignment(horizontal='center', vertical='center')
    column_alp = chr(ord('A') + worksheet.max_column - 1)
    range_end_cell = column_alp + str(worksheet.max_row)
    range_cell = 'A1:' + range_end_cell
    # print(f'全表单元格范围: {range_cell}')
    for row_cell in worksheet[range_cell]:
        for cell in row_cell:
            cell.alignment = alignment_center
            cell.border = border

    # 保存表格
    workbook.save(excel_file)
    workbook.close()
    print(f'美化表格保存路径: {excel_file}')


def crawling_goats():
    header = {
        "User-Agent": "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/"
                      "86.0.4240.198 Safari/537.36"
    }

    today = time.strftime('%Y-%m-%d')
    tm_rng = pd.date_range(end=today, periods=30, freq='D')
    print(f'tm_rng: {tm_rng}')

    all_data = []
    for tms in tm_rng:
        tm = tms.strftime('%Y-%m-%d')
        url = 'https://trade.500.com/jczq/?playid=270&g=2&date={}'.format(tm)
        # print(f'tm: {tm}')
        print(f'url: {url}')

        respone = requests.get(url, headers=header)
        content = respone.content.decode('ISO-8859-1')
        soup = BeautifulSoup(respone.text, 'lxml')
        trs = soup.find_all('tr', class_="bet-tb-tr bet-tb-end")

        for tr in trs:
            one_match_data = []
            # 编号
            td_no = tr.find_all('td', class_="td td-no")
            serial_number = td_no[0].text
            # print(f'serial_number: {serial_number}')
            one_match_data.append(serial_number)

            # 依次为主队排名':'','主队':'','比分':'','客队':'','客队排名': ''
            td_team = tr.find_all('td', class_="td td-team")
            match_detail = td_team[0].find_all('a')
            if len(match_detail) == 3:
                # print(f'主队: {match_detail[0].text}')
                # print(f'比分: {match_detail[1].text}')
                # print(f'客队: {match_detail[2].text}')
                one_match_data.append(match_detail[0].text)
                one_match_data.append(match_detail[1].text)
                one_match_data.append(match_detail[2].text)
            if len(match_detail) < 3:
                one_match_data.append(match_detail[0].text)
                one_match_data.append('VS')
                one_match_data.append(match_detail[1].text)

            # 进球数
            td_betbtn = tr.find_all('td', class_="td td-betbtn")
            p = td_betbtn[0].find_all('p')
            for match in p:
                # print(f'goat: {match.text}')
                one_match_data.append(match.text)

            all_data.append(one_match_data)

    print(f'all_data: {all_data}')
    df = pd.DataFrame(all_data,
                      columns=['serial_number', 'home_team', 'score', 'away_team', '0', '1', '2', '3', '4', '5', '6',
                               '7+'])
    # print(df)

    # 新建一个空白 DataFrame 用于存储结果
    result_df = pd.DataFrame()
    for index, row in df.iterrows():
        # 将当前行添加到结果 DataFrame 中
        result_df = result_df.append(row, ignore_index=True)
        # 添加一行空白行（用NaN填充）
        result_df = result_df.append(pd.Series([None] * len(df.columns)), ignore_index=True)

    # 保存新文件
    output_xlsx_file = 'goats-3.xlsx'
    result_df.to_excel(output_xlsx_file, index=False)

    # 美化excel
    format_excel(output_xlsx_file)


if __name__ == '__main__':
    crawling_goats()

