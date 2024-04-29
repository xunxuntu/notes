# -*- coding: utf-8 -*-
"""
    @File : go_for_shangan.py
    @Time : 2024-04-29
    Description : 
"""
import openpyxl
import pandas as pd
from datetime import datetime
from pathlib import Path
from tqdm import tqdm
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side


class ShangAn(object):
    """
        Make SHANGAN Be True
    """
    def __init__(self, matches_file, processed_file):
        self.xlsx_file = matches_file
        self.processed_file = processed_file

    def read_xlsx(self):
        """
            read matches data xlsx file
        :return:
        """
        df = pd.read_excel(self.xlsx_file)
        return df

    def double_tie_matches(self):
        """
            got double tie matches
        :return:
        """
        current_date = datetime.now().date()
        current_date_path = Path(str(current_date))
        current_date_path.mkdir(exist_ok=True)
        df = self.read_xlsx()
        output_df = pd.DataFrame(columns=df.columns)

        num_range = range(0, len(df), 2)  # 适用于每行两个球队比赛的情况
        for num in tqdm(num_range, desc='double_tie_matches Processing'):
            score = df.at[num, '比分']
            if pd.isna(score):
                continue  # 跳过NaN值的处理
            if score == 'VS':
                continue
            home_score, away_score = map(int, str(score).split(':'))

            # 让球
            handicap_score = df.at[num + 1, '让球']
            if isinstance(handicap_score, str) and len(handicap_score) > 4:  # 过滤出 " 单关-3" 这样的字符
                if handicap_score[:3] == " 单关":
                    handicap_score = int(handicap_score[3:])
            else:
                handicap_score = int(handicap_score)

            if home_score + handicap_score == away_score or home_score == away_score:
                # 双平局
                # output_df = output_df.append(df.iloc[num:num + 2]) # low version pandas use it
                output_df = pd.concat([output_df, df.iloc[num:num + 2]])

        match_count = output_df['比分'].count()
        all_match_count = df['比分'].count()
        percentage = match_count / all_match_count * 100
        print(f"双平场次: {match_count}, 占比: {percentage:.1f}%", )

        # save processed file
        output_df.to_excel(self.processed_file, index=False)

        # beautify processed file
        self.beautify_excel(self.processed_file)

    def low_sp_matches(self):
        """
            查找胜 or 负 sp在1.4以下的比赛
        :return:
        """
        xlsx_file = r'all_games.xlsx'
        df = pd.read_excel(xlsx_file, sheet_name='all')

        # 筛选 胜or负 sp 值低于1.4的场次
        threshold_num = 1.4
        special_sp_df = pd.DataFrame()
        for i in range(1, len(df), 2):
            # print(i)
            # print(df.loc[i-1, '负'])
            # print(type(df.loc[i, '负']))
            if df.loc[i - 1, '负'] == '未开售':
                continue
            if df.loc[i - 1, '负'] <= threshold_num or df.loc[i - 1, '胜'] <= threshold_num:
                special_sp_df = special_sp_df.append([df.loc[i - 1], df.loc[i]], ignore_index=True)

        # 导出为 excel
        special_sp_excel = '胜or负sp小于1.4场次.xlsx'
        special_sp_df.to_excel(special_sp_excel, index=False)

        # 美化 excel
        format_excel(special_sp_excel)

        # sp 正反路标记统计
        mark_pos_neg_outcome(xlsx_path=special_sp_excel,
                             sheet_name="Sheet1",
                             output_xlsx_file=special_sp_excel)


    def beautify_excel(self, excel_file):
        """
            make xlsx file more beautiful
        :return:
        """
        workbook = openpyxl.load_workbook(excel_file)
        sheets = workbook.sheetnames
        worksheet = workbook[sheets[0]]  # 读取第一个sheet表格

        # print(f'sheet的最大行数: {worksheet.max_row}')
        # print(f'sheet的最大列数: {worksheet.max_column}')

        # 合并单元格
        columns_list = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']  # 需要合并单元格修改的列
        for col in columns_list:
            for i in range(2, worksheet.max_row + 1, 2):
                start_cell = col + str(i)
                end_cell = col + str(i + 1)
                range_string = start_cell + ':' + end_cell
                worksheet.merge_cells(range_string)

        # 标记赛果 sp
        font = Font(color='FF0000', bold=True)
        for num in range(2, worksheet.max_row, 2):
            score_cell = 'G' + str(num)
            win_cell = 'K' + str(num)
            ping_cell = 'L' + str(num)
            lose_cell = 'M' + str(num)
            score = worksheet[score_cell].value
            if score == 'VS':
                continue
            home_score = int(score.split(':')[0])
            away_score = int(score.split(':')[1])

            # 让球
            handicap_score_cell = 'J' + str(num + 1)
            handicap_win_cell = 'K' + str(num + 1)
            handicap_ping_cell = 'L' + str(num + 1)
            handicap_lose_cell = 'M' + str(num + 1)
            handicap_score = str(worksheet[handicap_score_cell].value)
            if len(handicap_score) > 4:  # 过滤出 " 单关-3" 这样的字符
                if handicap_score[:3] == " 单关":
                    handicap_score = int(handicap_score[3:])
            else:
                handicap_score = int(handicap_score)

            if home_score > away_score:
                # 标记胜平负
                if worksheet[win_cell].value == '未开售':
                    pass
                else:
                    worksheet[win_cell].font = font
                if home_score + handicap_score > away_score:
                    worksheet[handicap_win_cell].font = font  # 标记让胜
                elif home_score + handicap_score == away_score:
                    worksheet[handicap_ping_cell].font = font  # 标记让平
                else:
                    worksheet[handicap_lose_cell].font = font  # 标记让负

            elif home_score == away_score:
                # 标记胜平负
                worksheet[ping_cell].font = font
                if home_score + handicap_score > away_score:
                    worksheet[handicap_win_cell].font = font  # 标记让胜
                elif home_score + handicap_score == away_score:
                    worksheet[handicap_ping_cell].font = font  # 标记让平
                else:
                    worksheet[handicap_lose_cell].font = font  # 标记让负

            else:
                # 标记胜平负
                worksheet[lose_cell].font = font
                if home_score + handicap_score > away_score:
                    worksheet[handicap_win_cell].font = font  # 标记让胜
                elif home_score + handicap_score == away_score:
                    worksheet[handicap_ping_cell].font = font  # 标记让平
                else:
                    worksheet[handicap_lose_cell].font = font  # 标记让负

        # 居中单元格, 绘制边框线
        # 创建边框线对象
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
        # 居中单元格
        alignment_center = Alignment(horizontal='center', vertical='center')
        column_alp = chr(ord('A') + worksheet.max_column - 1)
        range_end_cell = column_alp + str(worksheet.max_row)
        range_cell = 'A1:' + range_end_cell
        # print(f'全表单元格范围: {range_cell}')
        for row_cell in worksheet[range_cell]:
            for cell in row_cell:
                cell.alignment = alignment_center
                cell.border = border

        # 保存表格
        workbook.save(self.processed_file)
        workbook.close()
        print(f'美化表格保存路径: {self.processed_file}')


if __name__ == '__main__':
    input_file = "../data/matches_data.xlsx"
    output_file = "../result_sum/double_tie_matches.xlsx"
    shang_an = ShangAn(matches_file=input_file,
                       processed_file=output_file)
    shang_an.double_tie_matches()
